import openpyxl
import pandas as pd


def extract_xero_data(xero_file_path):
    xero_df = pd.read_excel(xero_file_path, skiprows=4)  # Skip the first 4 rows
    xero_df = xero_df.dropna(subset=['Account'])  # Drop rows where 'Account' is NaN
    xero_df = xero_df[['Account', '2023']]  # Select relevant columns
    xero_df['Account Code'] = xero_df['Account'].str.extract('(\d{4,})')  # Extract account codes
    xero_df = xero_df.dropna(subset=['Account Code'])  # Drop rows where 'Account Code' is NaN
    xero_df['Account Code'] = xero_df['Account Code'].astype(str).apply(lambda x: x.ljust(8, '0'))  # Pad account codes to 8 digits
    return xero_df

def map_to_excel(data, mapping, excel_path, target_sheet_name):
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook[target_sheet_name]

    # Track missing accounts
    missing_accounts = []
    found_accounts = []
    # Scan through columns F to M to find account codes and update column O
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=6, max_col=13):  # F=6, M=13
        for cell in row:
            account_code = cell.value
            if account_code and str(account_code).isdigit():
                account_code_str = str(account_code).ljust(8, '0')  # Pad to 8 digits
                mapped_account_code_str = mapping.get(account_code_str, account_code_str)
                if mapped_account_code_str in data['Account Code'].values:
                    print(f"Found account code {account_code_str} in the mapping")
                    row_idx = cell.row
                    value = data.loc[data['Account Code'] == mapped_account_code_str, '2023'].values[0]
                    found_accounts.append(account_code_str)
                    # first display the value in the cell
                    print(f"Row {row_idx}, Column {cell.column}: {value}")
                    sheet.cell(row=row_idx, column=19).value = value  # S=17
                else:
                    print(f"Account code {account_code_str} not found in the mapping")
                    missing_accounts.append(account_code_str)
    diff = set(data['Account Code'].values) - set(found_accounts)
    print(f"Accounts not found in the destination Excel file: {diff}")
    # Print missing accounts
    if missing_accounts:
        print("Accounts not found in the destination Excel file:")
        for account in set(missing_accounts):
            print(account)

    workbook.save(excel_path)

def main():
    # Hardcoded mapping dictionary
    mapping = {
        '42173000': '6187',
        '60330000': '6033',
        '60814000': '6138',
        '7454000': '7444',
        '60315000': '608112',
        '60340000': '60816',
        '61334000': '61338',
        '61338000': '61338',
        '61851000': '60813',
        '61854000': '6032'
    }

    # Extract data from Xero profit and loss report
    xero_file_path = "/Users/mehdi/Downloads/compta 2023/SUR_lu_SARL_-_Profit_and_Loss.xlsx"
    xero_data = extract_xero_data(xero_file_path)

    # Map and export the data to an existing Excel file in the "F3 CHARGES" sheet
    excel_file_path = "/Users/mehdi/Downloads/CNS/tempxxx/302497_1.xlsx"
    target_sheet_name = "F3 CHARGES"
    map_to_excel(xero_data, mapping, excel_file_path, target_sheet_name)


if __name__ == "__main__":
    main()
