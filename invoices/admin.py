import base64
import calendar
import csv
import datetime
import io
import os
from collections import defaultdict
from datetime import datetime as dt
from decimal import Decimal
from zoneinfo import ZoneInfo

import openpyxl
from PIL import Image
from admin_object_actions.admin import ModelAdminObjectActionsMixin
from constance import config
from django.contrib import admin
from django.contrib.admin import TabularInline
from django.contrib.admin.views.main import ChangeList
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin, csrf_protect_m
from django.contrib.auth.models import User
from django.core.cache import cache
from django.core.checks import messages
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q, Count
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse
from django.utils import timezone
from django.utils.html import format_html
from django.utils.http import urlencode
from django.utils.safestring import mark_safe
from django.utils.translation import gettext_lazy as _
from django_csv_exports.admin import CSVExportAdmin
from openpyxl.styles import PatternFill, Font

from dependence.invoicing import LongTermCareInvoiceFile, LongTermCareInvoiceItem
from dependence.longtermcareitem import LongTermPackage
from helpers.timesheet import build_use_case_objects
from invoices.action import export_to_pdf, set_invoice_as_sent, set_invoice_as_paid, set_invoice_as_not_paid, \
    set_invoice_as_not_sent, find_all_invoice_items_with_broken_file, \
    find_all_medical_prescriptions_and_merge_them_in_one_file, link_invoice_to_invoice_batch, create_google_contact, \
    cleanup_contacts, cleanup_some_contacts
from invoices.action_private import pdf_private_invoice
from invoices.action_private_participation import pdf_private_invoice_pp
from invoices.actions.carecodes import update_prices_for_september_2023
from invoices.actions.certificates import generate_pdf
from invoices.actions.invoices import generer_forfait_aev_juin_2024
# from invoices.actions.maps import calculate_distance_matrix
from invoices.actions.print_pdf import do_it, PdfActionType
from invoices.distancematrix import DistanceMatrix
from invoices.employee import Employee, EmployeeContractDetail, JobPosition, EmployeeAdminFile, EmployeeProxy
from invoices.enums.event import EventTypeEnum
from invoices.enums.holidays import HolidayRequestWorkflowStatus
from invoices.events import EventType, Event, AssignedAdditionalEmployee, ReportPicture, \
    create_or_update_google_calendar, EventList, EventGenericLink, EventLinkToMedicalCareSummaryPerPatientDetail, \
    EventLinkToCareCode, GenericTaskDescription
from invoices.filters.HolidayRequestFilters import FilteringYears, FilteringMonths
from invoices.filters.SmartEmployeeFilter import SmartEmployeeFilter, SmartUserFilterForVisits, SmartPatientFilter, \
    SmartMedicalPrescriptionFilter, DistanceMatrixSmartPatientFilter, IsInvolvedInHealthCareFilter, \
    SmarPatientFilterForVisits
from invoices.filters.SmartPatientFilter import UnderAssuranceDependanceFilter, UnderHeatWaveRiskFilter, \
    IsPatientDeceasedFilter
from invoices.forms import ValidityDateFormSet, HospitalizationFormSet, \
    PrestationInlineFormSet, PatientForm, SimplifiedTimesheetForm, SimplifiedTimesheetDetailForm, EventForm, \
    InvoiceItemForm, MedicalPrescriptionForm, AlternateAddressFormSet, EventLinkToMedicalCareSummaryPerPatientDetailForm
from invoices.gcalendar2 import PrestationGoogleCalendarSurLu
from invoices.googlemessages import post_webhook
from invoices.holidays import HolidayRequest, AbsenceRequestFile
from invoices.models import CareCode, Prestation, Patient, InvoiceItem, Physician, ValidityDate, MedicalPrescription, \
    Hospitalization, InvoiceItemBatch, InvoiceItemEmailLog, PatientAdminFile, InvoiceItemPrescriptionsList, \
    AlternateAddress, Alert, Bedsore, BedsoreEvaluation, BedsoreRiskAssessment, SubContractor, SubContractorAdminFile, \
    PatientSubContractorRelationship, ClientPatientRelationship
from invoices.modelspackage import InvoicingDetails
from invoices.notifications import notify_holiday_request_validation
from invoices.prefac import generate_flat_file, generate_flat_file_for_control
from invoices.resources import ExpenseCard, Car, MaintenanceFile, ConvadisOAuth2Token, CarBooking
from invoices.salaries import EmployeesMonthlyPayslipFile, EmployeePaySlip
from invoices.timesheet import Timesheet, TimesheetDetail, TimesheetTask, \
    SimplifiedTimesheetDetail, SimplifiedTimesheet, PublicHolidayCalendarDetail, PublicHolidayCalendar
from invoices.utils import EventCalendar
from invoices.visitmodels import EmployeeVisit
from invoices.xeromodels import XeroToken


@admin.register(CarBooking)
class CarBookingAdmin(admin.ModelAdmin):
    list_display = ('booking_date',)


@admin.register(JobPosition)
class JobPositionAdmin(admin.ModelAdmin):
    list_display = ('name', 'description', 'is_involved_in_health_care')


@admin.register(TimesheetTask)
class TimesheetTaskAdmin(admin.ModelAdmin):
    list_display = ('name', 'description')


# Define an inline admin descriptor for Employee model
# which acts a bit like a singleton
class EmployeeInline(admin.StackedInline):
    model = Employee
    can_delete = False
    verbose_name_plural = 'employee'
    extra = 0


class UserAdmin(BaseUserAdmin):
    inlines = (EmployeeInline,)
    actions = ['deactivate_user']

    def deactivate_user(self, request, queryset):
        # only super-users can do this
        if not request.user.is_superuser:
            self.message_user(request, "Vous n'êtes pas autorisé à effectuer cette action.",
                              level=messages.ERROR)
            return
        users_deactivated = []
        for user in queryset:
            user.is_active = False
            user.is_staff = False
            user.save()
            users_deactivated.append(user.username)
        self.message_user(request, "Utilisateurs désactivés: %s" % users_deactivated,
                          level=messages.INFO)


# Re-register UserAdmin
admin.site.unregister(User)
admin.site.register(User, UserAdmin)


class ValidityDateInline(admin.TabularInline):
    extra = 0
    model = ValidityDate
    formset = ValidityDateFormSet
    fields = ('start_date', 'end_date', 'gross_amount')
    search_fields = ['start_date', 'end_date', 'gross_amount']


@admin.register(CareCode)
class CareCodeAdmin(admin.ModelAdmin):
    list_display = ('code', 'name', 'reimbursed', 'latest_price')
    search_fields = ['code', 'name']
    inlines = [ValidityDateInline]
    actions = [update_prices_for_september_2023]
    readonly_fields = ('latest_price',)
    # actions = [update_prices_for_april_2022]


class EmployeeContractDetailInline(TabularInline):
    extra = 0
    model = EmployeeContractDetail


class EmployeeAdminFileInline(TabularInline):
    extra = 0
    model = EmployeeAdminFile


@admin.register(ClientPatientRelationship)
class ClientPatientRelationshipAdmin(admin.ModelAdmin):
    list_display = ('user', 'patient')
    autocomplete_fields = ['user', 'patient']


@admin.register(Employee)
class EmployeeAdmin(admin.ModelAdmin):
    inlines = [EmployeeContractDetailInline, EmployeeAdminFileInline]
    list_display = ('user', 'phone_number', 'start_contract', 'end_contract', 'occupation', 'abbreviation',
                    'employee_fte',)
    search_fields = ['user__last_name', 'user__first_name', 'user__email']
    readonly_fields = ['total_number_of_un_validated_events', 'minified_avatar', 'minified_avatar_base64']
    list_filter = [IsInvolvedInHealthCareFilter]
    actions = ['send_tomorrows_events', 'send_todays_events', 'work_certificate', 'contracts_situation_certificate', 'entry_declaration',
               'export_employees_data_to_csv',
               create_google_contact, cleanup_contacts, cleanup_some_contacts,
               'generate_annual_report_for_2023']

    def send_tomorrows_events(self, request, queryset):
        # if not super user, return
        if not request.user.is_superuser:
            self.message_user(request, "Vous n'êtes pas autorisé à effectuer cette action.",
                              level=messages.ERROR)
            return
        # get all the events for tomorrow and send them by email to employee
        tomorrow = timezone.now() + datetime.timedelta(days=1)
        # send the email
        employees_who_will_receive_email = []
        for emp in queryset:
            tomorrow_events = Event.objects.filter(day=tomorrow, employees=emp, state__in=[1, 2]).order_by(
                'time_start_event')
            # create a beautiful text in french that lists all events of the day
            text = ""
            if len(tomorrow_events) == 0:
                text = "Pas d'événements pour demain."
            for event in tomorrow_events:
                event_notes = event.notes if event.notes else "Soins habituels"
                text += f"🔘 de approx. {event.time_start_event} à approx. {event.time_end_event} chez {event.patient}: {event_notes} \n"
            emp.send_email_with_events(text, tomorrow)
            employees_who_will_receive_email.append(emp)
        self.message_user(request, "Email envoyé à %s employés. %s " % (len(employees_who_will_receive_email),
                                                                        employees_who_will_receive_email),
                          level=messages.INFO)
    def send_todays_events(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request, "Vous n'êtes pas autorisé à effectuer cette action.",
                              level=messages.ERROR)
            return
        today = timezone.now()
        employees_who_will_receive_email = []
        for emp in queryset:
            today_events = Event.objects.filter(day=today, employees=emp, state__in=[1, 2]).order_by(
                'time_start_event')
            text = ""
            if len(today_events) == 0:
                text = "Pas d'événements pour aujourd'hui."
            for event in today_events:
                event_notes = event.notes if event.notes else "Soins habituels"
                text += f"🔘 de approx. {event.time_start_event} à approx. {event.time_end_event} chez {event.patient}: {event_notes} \n"
            emp.send_email_with_events(text, today)
            employees_who_will_receive_email.append(emp)
        self.message_user(request, "Email envoyé à %s employés. %s " % (len(employees_who_will_receive_email),
                                                                        employees_who_will_receive_email),
                          level=messages.INFO)

    def save_model(self, request, obj, form, change):
        if 'avatar' in form.changed_data:
            original_image = form.cleaned_data['avatar']

            # Open the image using Pillow
            img = Image.open(original_image)

            # Resize/minify the image using LANCZOS resampling filter
            img.thumbnail((48, 48), Image.Resampling.LANCZOS)

            # Save the minified image to a BytesIO object
            in_memory_image = io.BytesIO()
            img_format = 'JPEG' if img.format == 'JPEG' else 'PNG'  # Adjust based on your needs
            img.save(in_memory_image, format=img_format)

            # Reset the position to the start of the byte array
            in_memory_image.seek(0)

            # Encode the byte array in base64
            encoded_avatar = base64.b64encode(in_memory_image.read()).decode()

            # Save the base64 encoded image to the minified_avatar_svg field
            obj.minified_avatar_base64 = f"data:image/{img_format.lower()};base64,{encoded_avatar}"

        super().save_model(request, obj, form, change)

    def generate_annual_report_for_2023(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request, "Vous n'êtes pas autorisé à effectuer cette action.",
                              level=messages.ERROR)
            return
        unique_employees = []
        # group by nationality and count
        nationality_counts = queryset.values('citizenship').annotate(count=Count('citizenship')).order_by('count')
        # generate a csv file with number of employees for year 2023
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="employees_2023.csv"'
        writer = csv.writer(response)
        writer.writerow(['Nationalité', 'Nombre d\'employés'])
        for nc in nationality_counts:
            writer.writerow([nc['citizenship'], nc['count']])
        city_counts = queryset.values('address').annotate(count=Count('address')).order_by('count')
        writer.writerow(['Ville', 'Nombre d\'employés'])
        for cc in city_counts:
            writer.writerow([cc['address'], cc['count']])
        age_group_dict = {}
        for e in queryset:
            age_group = e.age_group
            if age_group in age_group_dict:
                age_group_dict[age_group] += 1
            else:
                age_group_dict[age_group] = 1
        writer.writerow(['Groupe d\'age', 'Nombre d\'employés'])
        for key, value in age_group_dict.items():
            writer.writerow([key, value])

        all_events = Event.objects.filter(day__year=2023).filter(patient__is_under_dependence_insurance=True).exclude(
            patient__id=751).exclude(event_type_enum=EventTypeEnum.BIRTHDAY).exclude(
            event_type_enum=EventTypeEnum.GENERIC).exclude(event_type_enum=EventTypeEnum.BIRTHDAY)
        keep_only_events_after_1_mars_2023 = []
        for event in all_events:
            if event.day > datetime.date(2023, 3, 1):
                keep_only_events_after_1_mars_2023.append(event)
        writer.writerow(["Nombre d'événements pour les patients sous assurance dépendance",
                         len(keep_only_events_after_1_mars_2023)])

        return response

    # if not super user display only active employees
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.filter(end_contract__isnull=True)

    # property that display full-time equivalent of employee
    def display_fte(self, obj):
        return obj.calculate_fte()

    def entry_declaration(self, request, queryset):
        counter = 1
        file_data = ""
        for emp in queryset:
            if emp.end_contract:
                pass
            else:
                _last_first_name = "%s %s" % (emp.user.last_name.upper(), emp.user.first_name.capitalize())
                _matricule = "Matricule CNS: %s " % emp.sn_code
                _occupation = "Occupation: %s" % emp.occupation
                _address = "Demeurant au: %s" % emp.address
                # format all dates to dd/mm/yyyy
                _date_entree = "Date Entrée: %s " % emp.start_contract.strftime("%d/%m/%Y")
                # format date to dd/mm/yyyy
                if emp.virtual_career_anniversary_date:
                    _career_anniversary = "Anniversaire de carrière virtuelle: %s" % emp.virtual_career_anniversary_date.strftime(
                        "%d/%m/%Y")
                else:
                    _career_anniversary = "Anniversaire de carrière virtuelle: %s" % "Non défini"
                _citizenship = "Nationalité: %s" % emp.citizenship
                cd = EmployeeContractDetail.objects.filter(employee_link_id=emp.id, end_date__isnull=True).first()
                _contract_situation = "Contrat %s %s h. / semaine - salaire: %s / mois" % (
                    cd.contract_type, cd.number_of_hours, cd.monthly_wage)
                if emp.end_trial_period:
                    _trial_period = "Date fin période d'essai: %s" % emp.end_trial_period.strftime("%d/%m/%Y")
                else:
                    _trial_period = "Date fin période d'essai: %s" % "Non définie"
                _career_rank = "Grade: %s indice %s" % (cd.career_rank, cd.index)
                _bank_account_details = "Numéro de compte bancaire: %s" % emp.bank_account_number
                file_data += f"{_last_first_name} \n {_matricule} \n {_occupation} \n {_citizenship} \n {_address} \n " \
                             f"{_date_entree} {_career_anniversary}\n {_career_rank} \n {_contract_situation} \n " \
                             f"{_trial_period} \n " \
                             f"{_bank_account_details} \n"

            counter += 1
        response = HttpResponse(file_data, content_type='application/text charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="declaration_entree.txt"'
        return response

    def work_certificate(self, request, queryset):
        try:
            return generate_pdf(queryset)
        except ValidationError as ve:
            self.message_user(request, ve.message,
                              level=messages.ERROR)

    def calculate_etp_for_all_employees_for_the_year_2020(self, request, queryset):
        # get all the employees
        employees = Employee.objects.all()
        # loop through all the employees that are not terminated in 2020 and calculate their ETP
        who_was_working_in_2020 = []
        for emp in employees:
            if emp.end_contract and emp.end_contract.year == 2020:
                continue
            if emp.start_contract.year > 2020:
                continue
            if emp.end_contract and emp.end_contract.year < 2020:
                continue
            who_was_working_in_2020.append(emp)
        # return a csv file with the results
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="etp_stats.csv"'
        writer = csv.writer(response)
        # I need employee name, occupation, start date, end date, number of hours in contract
        writer.writerow(
            ['ID', 'Nom employe', 'Occupation', 'Date debut', 'Date fin', 'Moy. Nombre d\'heures par semaine en 2020'])
        for emp in who_was_working_in_2020:
            average_hours_per_week = emp.get_average_hours_per_week(
                datetime.date(2020, 1, 1), datetime.date(2020, 12, 31))
            if emp.employeecontractdetail_set.filter(end_date__isnull=True).first() is not None:
                writer.writerow([emp.id, emp.user.last_name, emp.occupation, emp.start_contract, emp.end_contract,
                                 average_hours_per_week])
            else:
                writer.writerow([emp.id, emp.user.last_name, emp.occupation, emp.start_contract, emp.end_contract,
                                 average_hours_per_week])
        return response

    def contracts_situation_certificate(self, request, queryset):
        counter = 1
        file_data = ""
        for emp in queryset:
            if emp.end_contract:
                file_data += "%d - %s %s FIN DE CONTRAT LE: %s \n" % (counter, emp.user.last_name.upper(),
                                                                      emp.user.first_name,
                                                                      emp.end_contract.strftime("%d/%m/%Y"))
            else:
                file_data += "%d - %s %s Occupation: %s Temps de travail: %s\n" % (counter,
                                                                                   emp.user.last_name.upper(),
                                                                                   emp.user.first_name,
                                                                                   emp.occupation,
                                                                                   emp.employeecontractdetail_set.filter(
                                                                                       end_date__isnull=True)
                                                                                   .first())
            counter += 1
        response = HttpResponse(file_data, content_type='application/text charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="contract_situation.txt"'
        return response

    work_certificate.short_description = "Certificat de travail"
    contracts_situation_certificate.short_description = "Situation des contrats"

    def export_employees_data_to_csv(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request, "Vous n'êtes pas autorisé à effectuer cette action.",
                              level=messages.ERROR)
            return
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="etp_stats.csv"'
        writer = csv.writer(response)
        _stats_date = datetime.date(2023, 4, 30)
        # _stats_date = datetime.date(2023, 7, 21)
        writer.writerow(['Identifiant anonyme', 'Année de naissance', 'Pays de résidence', 'Date début du contrat',
                         'Date fin du contrat (si connue)', 'CCT', 'Carrière', 'Echelon',
                         'Points au %s' % _stats_date.strftime("%d/%m/%Y"), 'Durée de travail hebdomadaire (en heures)',
                         "Structure d'affectation", 'Fonction'])
        for emp in queryset:
            # date 30/04/2023
            _employee_contract = emp.get_contrat_at_date(_stats_date)
            if not _employee_contract:
                continue
            _rank = _employee_contract.career_rank
            # _rank looks like this C5/120 we need to split it
            cct = "-"
            _echelon_rank = []
            if not _rank:
                _echelon_rank.append("NA")
            else:
                _echelon_rank = _rank.split("/")
            if len(_echelon_rank) == 1:
                _echelon_rank.append("NA")
                cct = "-"
            else:
                cct = "SAS"
            # _emp_start in french format date
            _emp_start = _employee_contract.start_date.strftime("%d/%m/%Y")
            _emp_end = _employee_contract.end_date.strftime("%d/%m/%Y") if _employee_contract.end_date else "-"
            writer.writerow([emp.id, emp.birth_date.year, emp.address, _emp_start, _emp_end, cct,
                             _echelon_rank[0], _echelon_rank[1], "", _employee_contract.number_of_hours,
                             "UNIQUE", emp.get_occupation()])
        return response

    # actions = [work_certificate, 'delete_in_google_calendar']
    def get_readonly_fields(self, request, obj=None):
        if request.user.is_superuser:
            # If the user is a superuser, no fields are read-only
            return []
        else:
            # If the user is not a superuser, all fields except 'field1' and 'field2' are read-only
            # user can only edit his own profile
            if obj is not None and obj.user == request.user:
                return [field.name for field in obj._meta.fields if field.name not in ['bio', 'to_be_published_on_www']]
            else:
                return [field.name for field in obj._meta.fields]

    def get_fieldsets(self, request, obj=None):
        if request.user.is_superuser or request.user == obj.user:
            return super().get_fieldsets(request, obj)
        else:
            return [
                (None, {'fields': [field.name for field in obj._meta.fields if field.name != 'bank_account_number']})]

    def delete_in_google_calendar(self, request, queryset):
        if not request.user.is_superuser:
            return
        counter = 0
        for e in queryset:
            calendar_gcalendar = PrestationGoogleCalendarSurLu()
            counter = calendar_gcalendar.delete_all_events_from_calendar(e.user.email)
        self.message_user(request, "%s évenements supprimés." % counter,
                          level=messages.INFO)


class ExpenseCardDetailInline(TabularInline):
    extra = 0
    model = ExpenseCard


class MaintenanceFileInline(TabularInline):
    extra = 0
    model = MaintenanceFile


class NewCarList(Car):
    class Meta:
        proxy = True
        verbose_name = "Véhicule *"
        verbose_name_plural = "Véhicules *"


@admin.register(NewCarList)
class NewCarListAdmin(admin.ModelAdmin):
    # change_list_template = 'car/new_car_list_admin.html'
    change_list_template = 'angular-car/app-car-list.html'

    # reference the js file call it cars.js
    class Media:
        js = ('car/cars.js',)

    # display only cars that are is_blue_link_connected True
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.filter(is_connected_to_bluelink=True)


@admin.register(Car)
class CarAdmin(admin.ModelAdmin):
    inlines = [ExpenseCardDetailInline, MaintenanceFileInline]
    list_display = (
        'name', 'licence_plate', 'pin_codes', 'geo_localisation_of_car_url', 'battery_or_fuel_level', 'car_movement')

    def geo_localisation_of_car_url(self, obj):
        _geo_localisation_of_car = obj.geo_localisation_of_car
        if type(_geo_localisation_of_car) is not tuple and _geo_localisation_of_car.startswith(
                'n/a') or 'Error' in _geo_localisation_of_car:
            return _geo_localisation_of_car
        else:

            url = 'https://maps.google.com/?q=%s,%s' % (_geo_localisation_of_car[1],
                                                        _geo_localisation_of_car[2])
            address = obj.address
            return format_html("<a href='%s'>%s</a>" % (url, address))
        return _geo_localisation_of_car

    geo_localisation_of_car_url.allow_tags = True
    geo_localisation_of_car_url.short_description = "Dernière position connue"


class HospitalizationInline(admin.TabularInline):
    extra = 0
    model = Hospitalization
    formset = HospitalizationFormSet
    fields = ('start_date', 'end_date', 'description')


class PatientAdminFileInline(admin.TabularInline):
    extra = 0
    model = PatientAdminFile

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.order_by('-file_date')


class AlternateAddressInline(admin.TabularInline):
    model = AlternateAddress
    extra = 0
    formset = AlternateAddressFormSet


class MedicalPrescriptionInlineAdmin(admin.TabularInline):
    extra = 0
    model = MedicalPrescription
    readonly_fields = ('thumbnail_img',)
    autocomplete_fields = ['prescriptor']
    fields = [field.name for field in MedicalPrescription._meta.fields if field.name not in ["id", "file"]]
    ordering = ['-date']

    def scan_preview(self, obj):
        return obj.image_preview

    scan_preview.allow_tags = True


class BedsoreRiskAssessment(admin.TabularInline):
    model = BedsoreRiskAssessment
    extra = 0


class SubContractorAdminFileInline(admin.TabularInline):
    model = SubContractorAdminFile
    extra = 0

    # all the fields should be readonly if not superuser
    def get_readonly_fields(self, request, obj=None):
        if request.user.is_superuser:
            return []
        return [field.name for field in self.model._meta.fields]


@admin.register(SubContractor)
class SubContractorAdmin(admin.ModelAdmin):
    list_display = ('name', 'phone_number', 'provider_code')
    search_fields = ['name', 'provider_code', 'phone_number']
    inlines = [SubContractorAdminFileInline]
    actions = ['create_subcontractor_in_xero']
    readonly_fields = ('created_by',)

    def create_subcontractor_in_xero(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request, "Vous n'êtes pas autorisé à effectuer cette action.",
                              level=messages.ERROR)
            return
        contractors_created = []
        for sub in queryset:
            contractors_created.append(sub.create_subcontractor_in_xero())
        self.message_user(request, "Contact(s) créé(s) dans Xero. %s" % contractors_created,
                          level=messages.INFO)

    # all the fields should be readonly if not superuser
    def get_readonly_fields(self, request, obj=None):
        if request.user.is_superuser:
            return []
        return [field.name for field in self.model._meta.fields]

    # can only view the subcontractors that were created by user, or is part of the same admin group
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
            # Get the groups of the request user
        user_groups = request.user.groups.all()
        user_group_names = [group.name for group in user_groups]

        # Filter the queryset to include SubContractors created by the request user
        # or by users who are in the same group as the request user
        return qs.filter(
            Q(created_by=request.user) | Q(created_by__groups__in=user_groups) | Q(name__in=user_group_names))

    def save_model(self, request, obj, form, change):
        if not obj.pk:  # if object is being created, automatically set created_by
            obj.created_by = request.user
        super().save_model(request, obj, form, change)


class SubContractorInline(admin.TabularInline):
    model = PatientSubContractorRelationship
    extra = 1


@admin.register(Patient)
class PatientAdmin(CSVExportAdmin):
    list_filter = ('is_under_dependence_insurance', UnderHeatWaveRiskFilter, IsPatientDeceasedFilter)
    list_display = ('name', 'first_name', 'phone_number', 'code_sn', 'participation_statutaire', 'age')
    csv_fields = ['name', 'first_name', 'address', 'zipcode', 'city',
                  'country', 'phone_number', 'email_address', 'date_of_death']
    readonly_fields = ('age', 'link_to_invoices', 'link_to_medical_prescriptions', 'link_to_events',
                       'link_to_all_events_of_last_3_months')
    search_fields = ['name', 'first_name', 'code_sn', 'zipcode', 'city', 'phone_number', 'email_address']
    # actions = [calculate_distance_matrix]
    form = PatientForm
    actions = [generer_forfait_aev_juin_2024, "generate_annual_report_for_2023",
               "number_of_hospitalizations_in_2023",
               "number_of_dead_patients_in_2023", "number_of_patients_in_2023", "filter_without_gender", ]
    inlines = [HospitalizationInline, MedicalPrescriptionInlineAdmin, PatientAdminFileInline, AlternateAddressInline,
               BedsoreRiskAssessment, SubContractorInline]

    def age_moyen(self, request, queryset):
        total_age = 0
        for patient in queryset:
            total_age += patient.age
        print("Age moyen des patients sélectionnés: %s" % (total_age / len(queryset)))

    def minima_age(self, request, queryset):
        minimum_age = 100
        patient_ids = []
        for patient in queryset:
            patient_ids.append(patient.id)
            if patient.age < minimum_age:
                minimum_age = patient.age
        print("*** Patient ids: %s" % patient_ids)
        print("*** Age minimum des patients sélectionnés: %s" % minimum_age)

    def maximum_age(self, request, queryset):
        minimum_age = 0
        patient_ids = []
        for patient in queryset:
            patient_ids.append(patient.id)
            if patient.age > minimum_age:
                minimum_age = patient.age
        print("*** Patient ids: %s" % patient_ids)
        print("Age minimum des patients sélectionnés: %s" % minimum_age)

    def number_of_hospitalizations_in_2023(self, request, queryset):
        total_hospitalizations_f = 0
        total_hospitalizations_h = 0
        patient_ids = [1356, 1325, 1313, 1309, 1210, 1070, 1034, 945, 864, 749, 482, 480, 331, 198]
        for patient_id in patient_ids:
            patient = Patient.objects.get(id=patient_id)
            if patient.gender == 'FEM':
                total_hospitalizations_f += Hospitalization.objects.filter(start_date__year=2023,
                                                                           patient=patient).count()
            elif patient.gender == 'MAL':
                total_hospitalizations_h += Hospitalization.objects.filter(start_date__year=2023,
                                                                           patient=patient).count()
            else:
                print("Patient %s has no gender" % patient)
        print("Nombre total d'hospitalisations  de femmes en 2023: %s" % total_hospitalizations_f)
        print("Nombre total d'hospitalisations  d'hommes en 2023: %s" % total_hospitalizations_h)

    def number_of_dead_patients_in_2023(self, request, queryset):
        patient_ids = [1356, 1325, 1313, 1309, 1210, 1070, 1034, 945, 864, 749, 482, 480, 331, 198]
        total_dead_patients_f = Patient.objects.filter(date_of_death__year=2023, gender='FEM',
                                                       id__in=patient_ids).count()
        total_dead_patients_h = Patient.objects.filter(date_of_death__year=2023, gender='MAL',
                                                       id__in=patient_ids).count()
        print("Nombre total de patients f décédés en 2023: %s" % total_dead_patients_f)
        print("Nombre total de patients h décédés en 2023: %s" % total_dead_patients_h)
        list_names_patients_alive = []
        for patient_id in patient_ids:
            patient = Patient.objects.get(id=patient_id)
            if not patient.date_of_death:
                list_names_patients_alive.append(patient.name)
        print("Patients encore en vie: %s" % list_names_patients_alive)

    def number_of_patients_in_2023(self, request, queryset):
        # Define the age groups
        age_groups = [
            {"name": "<60", "age_from": 0, "age_to": 59},
            {"name": "60-79", "age_from": 60, "age_to": 79},
            {"name": "80-99", "age_from": 80, "age_to": 99},
            {"name": ">=100", "age_from": 100, "age_to": None},
        ]

        # Initialize a dictionary to store the results
        results = {}

        events = Event.objects.filter(day__year=2023).filter(patient__is_under_dependence_insurance=True).exclude(
            patient__id=751).exclude(event_type_enum=EventTypeEnum.BIRTHDAY)
        # Group by patient and count distinct patients
        unique_patient_ids = events.values_list('patient', flat=True).distinct()
        unique_patients_f = Patient.objects.filter(id__in=unique_patient_ids, gender='FEM')
        unique_patients_h = Patient.objects.filter(id__in=unique_patient_ids, gender='MAL')

        print("Number of unique patients with at least one event in 2023: ", len(unique_patient_ids))
        print("Number of unique patients f with at least one event in 2023: ", len(unique_patients_f))
        print("Number of unique patients h with at least one event in 2023: ", len(unique_patients_h))

        # Initialize a dictionary to store the classification results
        classification = {group["name"]: [] for group in age_groups}
        for patient in unique_patients_f:
            age = patient.age
            # Classify the patient based on their age
            for group in age_groups:
                if (group["age_to"] is None and age >= group["age_from"]) or \
                        (group["age_from"] <= age <= group["age_to"]):
                    classification[group["name"]].append(patient)
                    break

        # Print the classification results
        for group in age_groups:
            # print(f"Number of patients in the age group {group['name']}: {len(classification[group['name']])}")
            print(f"Number of patients in the age group {group['name']}: {classification[group['name']]}")

    def filter_without_gender(self, request, queryset):
        queryset = queryset.filter(gender__isnull=True)
        self.message_user(request, "Filtered patients without gender set.")

    def generate_annual_report_for_2023(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request, "Vous n'êtes pas autorisé à effectuer cette action.",
                              level=messages.ERROR)
            return
        unique_patients = []
        unique_patients_under_assurance_dependance = []
        # generate a csv file with number of patients for year 2023
        for invoice_item in InvoiceItem.objects.filter(invoice_date__year=2023):
            if invoice_item.patient not in unique_patients:
                unique_patients.append(invoice_item.patient)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="patients_2023.csv"'
        writer = csv.writer(response)
        writer.writerow(['Nom', 'Prénom', 'Date de naissance', 'Adresse', 'Code postal', 'Ville', 'Pays', 'Téléphone',
                         'Email', 'Under assurance dependance'])
        for patient in unique_patients:
            writer.writerow([patient.name, patient.first_name, patient.birth_date, patient.address, patient.zipcode,
                             patient.city, patient.country, patient.phone_number, patient.email_address])
        for event in Event.objects.filter(day__year=2023):
            if event.patient and event.patient not in unique_patients and not event.patient.is_under_dependence_insurance:
                unique_patients.append(event.patient)
                if event.patient:
                    writer.writerow(
                        [event.patient.name, event.patient.first_name, event.patient.birth_date, event.patient.address,
                         event.patient.zipcode, event.patient.city, event.patient.country, event.patient.phone_number,
                         event.patient.email_address, "False"])
                elif event.event_type_enum == EventTypeEnum.BIRTHDAY:
                    print("Event %s has no patient" % event)
                else:
                    print("Event %s has no patient" % event)
            elif event.patient and event.patient not in unique_patients_under_assurance_dependance and event.patient.is_under_dependence_insurance:
                unique_patients_under_assurance_dependance.append(event.patient)
                if event.patient:
                    writer.writerow(
                        [event.patient.name, event.patient.first_name, event.patient.birth_date, event.patient.address,
                         event.patient.zipcode, event.patient.city, event.patient.country, event.patient.phone_number,
                         event.patient.email_address, "True"])
                elif event.event_type_enum == EventTypeEnum.BIRTHDAY:
                    print("Event %s has no patient" % event)
                else:
                    print("Event %s has no patient" % event)
        print("Number of unique patients %s" % len(unique_patients))
        print(
            "Number of unique patients under assurance dependance %s" % len(unique_patients_under_assurance_dependance))
        # write a second row containing the number of unique patients
        writer.writerow(["Nombre de patients", len(unique_patients)])
        writer.writerow(
            ["Nombre de patients sous assurance dépendance", len(unique_patients_under_assurance_dependance)])
        # now calculate the number of employees that were working in 2022
        employees = Employee.objects.all()
        who_was_working_in_2022 = []
        for emp in employees:
            if emp.end_contract and emp.end_contract.year < 2023:
                continue
            # check if not already in the list
            if emp not in who_was_working_in_2022:
                who_was_working_in_2022.append(emp)
        # append all employees that were working in 2022 with the name
        for emp in who_was_working_in_2022:
            writer.writerow(
                [emp.user.last_name, emp.user.first_name, emp.user.email, emp.start_contract, emp.end_contract])
        # write a third row containing the number of unique employees
        writer.writerow(["Nombre d'employés", len(who_was_working_in_2022)])
        return response

    def link_to_invoices(self, instance):
        url = f'{reverse("admin:invoices_invoiceitem_changelist")}?patient__id={instance.id}'
        return mark_safe('<a href="%s">%s</a>' % (url, "cliquez ici (%d)" % InvoiceItem.objects.filter(
            patient_id=instance.id).count()))

    link_to_invoices.short_description = "Factures client"

    def link_to_medical_prescriptions(self, instance):
        url = f'{reverse("admin:invoices_medicalprescription_changelist")}?patient__id={instance.id}'
        return mark_safe('<a href="%s">%s</a>' % (url, "cliquez ici (%d)" % MedicalPrescription.objects.filter(
            patient_id=instance.id).count()))

    def link_to_events(self, instance):
        url = f'{reverse("admin:invoices_eventlist_changelist")}?event_type_enum__exact=GENERIC&patient__id={instance.id}'
        return mark_safe(
            '<a href="%s">%s</a>' % (url, "Tous les événements Generic du patient ici (%d)" % Event.objects.filter(
                patient_id=instance.id, event_type_enum__exact=EventTypeEnum.GENERIC).count()))

    def link_to_all_events_of_last_3_months(self, instance):
        url = f'{reverse("admin:invoices_eventlist_changelist")}?patient__id={instance.id}&day__gte={datetime.date.today() - datetime.timedelta(days=90)}&day__lte={datetime.date.today()}&o=-1.-2'
        return mark_safe(
            '<a href="%s">%s</a>' % (url, "Deniers passages du patient (%d)" % Event.objects.filter(
                patient_id=instance.id, day__gte=datetime.date.today() - datetime.timedelta(days=90),
                day__lte=datetime.date.today()).count()))

    link_to_all_events_of_last_3_months.short_description = "Deniers passages du patient"
    link_to_medical_prescriptions.short_description = "Ordonnances client"
    link_to_events.short_description = "Evénements client"

    def has_csv_permission(self, request):
        """Only super users can export as CSV"""
        if request.user.is_superuser:
            return True


class BedsoreEvaluationInline(admin.TabularInline):
    model = BedsoreEvaluation
    extra = 1


@admin.register(Bedsore)
class BedsoreAdmin(admin.ModelAdmin):
    inlines = [BedsoreEvaluationInline]
    autocomplete_fields = ['patient']


# @admin.register(Prestation)
# class PrestationAdmin(admin.ModelAdmin):
#     # from invoices.invaction import create_invoice_for_health_insurance
#     def formfield_for_foreignkey(self, db_field, request, **kwargs):
#         if db_field.name == "employee":
#             kwargs["queryset"] = Employee.objects.filter(owner=request.user)
#         return super().formfield_for_foreignkey(db_field, request, **kwargs)
#
#     list_filter = ('invoice_item__patient', 'invoice_item', 'carecode')
#     date_hierarchy = 'date'
#     list_display = ('carecode', 'date')
#     search_fields = ['carecode__code', 'carecode__name']
#     # actions = [create_invoice_for_health_insurance]
#
#
#
#     def get_changeform_initial_data(self, request):
#         initial = {}
#         user = request.user
#         try:
#             employee = user.employee
#             initial['employee'] = employee.id
#         except ObjectDoesNotExist:
#             pass
#
#         return initial
#
#     def get_inline_formsets(self, request, formsets, inline_instances, obj=None):
#         initial = {}
#         print(initial)


@admin.register(Physician)
class PhysicianAdmin(admin.ModelAdmin):
    list_filter = ('city',)
    list_display = ('full_name_from_cns', 'name', 'phone_number', 'provider_code')
    search_fields = ['name', 'first_name', 'provider_code', 'phone_number', 'city', 'zipcode', 'full_name_from_cns']
    actions = ['cleanup_provider_codes']

    def cleanup_provider_codes(self, request, queryset):
        number_of_physicians = queryset.count()
        for physician in queryset:
            dirty_provider_code = physician.provider_code.strip()
            # remove anything that is not a digit
            physician.provider_code = ''.join(filter(str.isdigit, dirty_provider_code))
            physician.save()
        self.message_user(request, "%s codes nettoyés." % number_of_physicians,
                          level=messages.INFO)


# def migrate_from_g_to_cl(modeladmin, request, queryset):
#     ps = MedicalPrescription.objects.all()
#     for p in ps:
#         if p.file and p.file.url and not p.image_file:
#             print(p.file)
#             local_storage = FileSystemStorage()
#             newfile = ContentFile(p.file.read())
#             relative_path = local_storage.save(p.file.name, newfile)
#
#             print("relative path %s" % relative_path)
#             up = uploader.upload(local_storage.location + "/" + relative_path, folder="medical_prescriptions/")
#             p.image_file = up.get('public_id')
#             p.save()
#             # break


@admin.register(MedicalPrescription)
class MedicalPrescriptionAdmin(admin.ModelAdmin):
    date_hierarchy = 'date'
    list_filter = ('date', 'prescriptor', 'patient')
    # list_display = ('date', 'prescriptor', 'patient', 'link_to_invoices', 'image_preview')
    list_display = ('date', 'prescriptor', 'patient', 'link_to_invoices')
    fields = ('prescriptor', 'patient', 'date', 'end_date', 'notes', 'file_upload', 'thumbnail_img',)
    search_fields = ['date', 'prescriptor__name', 'prescriptor__first_name', 'patient__name', 'patient__first_name']
    autocomplete_fields = ['prescriptor', 'patient']
    exclude = ('file',)
    form = MedicalPrescriptionForm

    # list_per_page = 5

    # actions = [migrate_from_g_to_cl]
    def link_to_invoices(self, obj):
        count = obj.med_prescription_multi_invoice_items.count()
        url = (
                reverse("admin:invoices_invoiceitem_changelist")
                + "?"
                + urlencode({"prescriptions__medical_prescription__id": f"{obj.id}"})
        )
        return format_html('<a href="{}">{} facture(s)</a>', url, count)

    link_to_invoices.short_description = "Factures client"


class PrestationInline(TabularInline):
    extra = 0
    max_num = InvoiceItem.PRESTATION_LIMIT_MAX
    model = Prestation
    formset = PrestationInlineFormSet
    fields = ('carecode', 'date', 'quantity', 'employee', 'copy', 'delete')
    autocomplete_fields = ['carecode']
    readonly_fields = ('copy', 'delete',)
    search_fields = ['carecode', 'date', 'employee']
    ordering = ['date']
    can_order = True

    class Media:
        js = [
            'admin/js/vendor/jquery/jquery.min.js',
            'admin/js/jquery.init.js',
            "js/inline-copy.js",
            "js/inline-delete.js",
        ]
        css = {
            'all': ('css/inline-prestation.css',)
        }

    def copy(self, obj):
        return format_html("<a href='#' class='copy_inline'>Copy</a>")

    def delete(self, obj):
        url = reverse('delete-prestation')
        return format_html("<a href='%s' class='deletelink' data-prestation_id='%s'>Delete</a>" % (url, obj.id))

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "employee":
            kwargs["queryset"] = Employee.objects.all().order_by("-end_contract", "abbreviation")
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    copy.allow_tags = True
    delete.allow_tags = True


@admin.register(InvoicingDetails)
class InvoicingDetailsAdmin(admin.ModelAdmin):
    list_display = ('provider_code', 'name', 'default_invoicing')


class InvoiceItemPrescriptionsListInlines(TabularInline):
    model = InvoiceItemPrescriptionsList
    extra = 0
    fields = ('medical_prescription',)
    autocomplete_fields = ['medical_prescription']


@admin.register(InvoiceItem)
class InvoiceItemAdmin(admin.ModelAdmin):
    class Media:
        css = {
            "all": ("css/invoice_item.css",)
        }

    from invoices.action import export_to_pdf, export_to_pdf_with_medical_prescription_files
    from invoices.action_private import pdf_private_invoice
    from invoices.action_private_participation import pdf_private_invoice_pp
    from invoices.action_depinsurance import export_to_pdf2
    form = InvoiceItemForm
    date_hierarchy = 'invoice_date'
    list_display = ('invoice_number', 'patient', 'invoice_month', 'invoice_sent', 'invoice_paid',
                    'number_of_prestations', 'invoice_details', 'has_medical_prescription')
    list_filter = ['invoice_date', 'invoice_details', 'invoice_sent', 'invoice_paid', SmartPatientFilter,
                   SmartMedicalPrescriptionFilter, 'created_by', 'prescriptions__medical_prescription__id']
    search_fields = ['patient__name', 'patient__first_name', 'invoice_number', 'patient__code_sn']
    readonly_fields = ('medical_prescription_preview', 'created_at', 'updated_at', 'batch', 'xero_invoice_url')
    autocomplete_fields = ['patient']

    def has_medical_prescription(self, obj):
        return InvoiceItemPrescriptionsList.objects.filter(invoice_item=obj).exists()

    has_medical_prescription.boolean = True

    def cns_invoice_bis(self, request, queryset):
        try:
            return do_it(queryset, action=PdfActionType.CNS)
        except ValidationError as ve:
            self.message_user(request, ve.message,
                              level=messages.ERROR)

    cns_invoice_bis.short_description = "CNS Invoice (new)"

    def action_remove_invoice_from_batch(self, request, queryset):
        # only super-users can do this
        if not request.user.is_superuser:
            self.message_user(request, "Vous n'êtes pas autorisé à effectuer cette action.",
                              level=messages.ERROR)
            return
        for invoice_item in queryset:
            invoice_item.batch = None
            invoice_item.save()
        self.message_user(request, "Factures retirées du batch.",
                          level=messages.INFO)

    def pdf_private_invoice_pp_bis(self, request, queryset):
        try:
            return do_it(queryset, action=PdfActionType.PERSONAL_PARTICIPATION)
        except ValidationError as ve:
            self.message_user(request, ve.message,
                              level=messages.ERROR)

    pdf_private_invoice_pp_bis.short_description = "Facture client participation personnelle (new)"

    actions = [action_remove_invoice_from_batch, link_invoice_to_invoice_batch, generate_flat_file,
               find_all_medical_prescriptions_and_merge_them_in_one_file, find_all_invoice_items_with_broken_file,
               export_to_pdf, export_to_pdf_with_medical_prescription_files, pdf_private_invoice_pp,
               pdf_private_invoice, export_to_pdf2, cns_invoice_bis, pdf_private_invoice_pp_bis, set_invoice_as_sent,
               set_invoice_as_paid, set_invoice_as_not_paid, set_invoice_as_not_sent]
    inlines = [InvoiceItemPrescriptionsListInlines, PrestationInline]
    fieldsets = (
        (None, {
            'fields': ('invoice_number', 'is_private', 'patient', 'invoice_date', 'invoice_details')
        }),
        ('Advanced options', {
            'classes': ('collapse',),
            'fields': ('accident_id', 'accident_date', 'is_valid', 'validation_comment',
                       'patient_invoice_date', 'invoice_send_date', 'invoice_sent', 'invoice_paid',
                       'medical_prescription', 'subcontractor', 'created_at', 'updated_at', 'batch',
                       'xero_invoice_url'),
        }),
    )
    verbose_name = u"Mémoire d'honoraire"
    verbose_name_plural = u"Mémoires d'honoraire"

    def medical_prescription_preview(self, obj):
        return obj.medical_prescription.image_preview()

    medical_prescription_preview.allow_tags = True

    def changelist_view(self, request, extra_context=None):
        changelist = ChangeList(request,
                                self.model,
                                list(self.get_list_display(request)),
                                self.get_list_display_links(request, self.get_list_display(request)),
                                self.get_list_filter(request),
                                self.date_hierarchy,
                                self.get_search_fields(request),
                                self.list_select_related,
                                self.list_per_page,
                                self.list_max_show_all,
                                self.list_editable,
                                self,
                                self.sortable_by,
                                self.get_search_results)

        # Get a distinct list of patients for the filtered queryset
        patients = list(set(changelist.get_queryset(request).values_list('patient__id', 'patient__name')))
        prescription_ids = list(
            set(changelist.get_queryset(request).values_list('prescriptions__medical_prescription__id', flat=True)))

        medical_prescriptions = MedicalPrescription.objects.filter(id__in=prescription_ids)

        # Attach it to request
        request.dynamic_patient_choices = [(str(patient_id), patient_name) for patient_id, patient_name in patients]
        request.dynamic_medical_prescription_choices = [(str(prescription.id), str(prescription)) for prescription in
                                                        medical_prescriptions]

        return super().changelist_view(request, extra_context)

    def response_change(self, request, obj):
        queryset = InvoiceItem.objects.filter(id=obj.id)
        if "_print_cns" in request.POST:
            return export_to_pdf(self, request, queryset)
            # matching_names_except_this = self.get_queryset(request).filter(name=obj.name).exclude(pk=obj.id)
            # matching_names_except_this.delete()
            # obj.is_unique = True
            # obj.save()
            # self.message_user(request, "This villain is now unique")
            # return HttpResponseRedirect(".")
        if "_print_private_invoice" in request.POST:
            return pdf_private_invoice(self, request, queryset)
        if "_print_personal_participation" in request.POST:
            return pdf_private_invoice_pp(self, request, queryset)
        if "_email_private_invoice" in request.POST:
            if hasattr(queryset[0].patient, 'email_address'):
                if not queryset[0].patient.email_address:
                    self.message_user(request, "Le patient n'a pas d'adresse email définie.",
                                      level=messages.ERROR)
                    return HttpResponseRedirect(request.path)
                if pdf_private_invoice(self, request, queryset, attach_to_email=True):
                    self.message_user(request, "La facture a bien été envoyée au client.",
                                      level=messages.INFO)
                else:
                    self.message_user(request, "La facture n'a pas pu être envoyée au client.",
                                      level=messages.ERROR)
                return HttpResponseRedirect(request.path)
        elif "_email_personal_participation" in request.POST:
            if hasattr(queryset[0].patient, 'email_address'):
                if not queryset[0].patient.email_address:
                    self.message_user(request, "Le patient n'a pas d'adresse email définie.",
                                      level=messages.ERROR)
                    return HttpResponseRedirect(request.path)
            if pdf_private_invoice_pp(self, request, queryset, attach_to_email=True):
                self.message_user(request, "La facture a bien été envoyée au client.",
                                  level=messages.INFO)
            else:
                self.message_user(request, "La facture n'a pas pu être envoyée au client.",
                                  level=messages.ERROR)
            return HttpResponseRedirect(request.path)
        elif "_email_private_invoice_xero" in request.POST:
            if pdf_private_invoice(self, request, queryset, attach_to_email=True,
                                   only_to_xero_or_any_accounting_system=True):
                self.message_user(request, "La facture a bien été envoyée à Xero.",
                                  level=messages.INFO)
            else:
                self.message_user(request, "La facture n'a pas pu être envoyée à Xero.",
                                  level=messages.ERROR)
            return HttpResponseRedirect(request.path)
        elif "_email_personal_participation_xero" in request.POST:
            if pdf_private_invoice_pp(self, request, queryset, attach_to_email=True,
                                      only_to_xero_or_any_accounting_system=True):
                self.message_user(request, "La facture a bien été à Xero.",
                                  level=messages.INFO)
            else:
                self.message_user(request, "La facture n'a pas pu être envoyée à Xero.",
                                  level=messages.ERROR)
            return HttpResponseRedirect(request.path)
        return HttpResponseRedirect(request.path)


class InvoiceItemInlineAdmin(admin.TabularInline):
    show_change_link = True
    max_num = 0
    extra = 0
    model = InvoiceItem
    readonly_fields = ('invoice_number', 'invoice_date',)
    fields = ('invoice_number', 'invoice_date',)
    ordering = ('invoice_date',)
    can_delete = False


@admin.register(InvoiceItemBatch)
class InvoiceItemBatchAdmin(ModelAdminObjectActionsMixin, admin.ModelAdmin):
    inlines = [InvoiceItemInlineAdmin]
    readonly_fields = ('count_invoices', 'created_date', 'modified_date')
    list_filter = ('start_date', 'end_date', 'batch_type', 'batch_description')
    list_display = ('start_date', 'end_date', 'send_date', 'count_invoices', 'batch_type', 'batch_description',
                    'display_object_actions_list')
    actions = [generate_flat_file_for_control, 'merge_invoices']

    object_actions = [
        {
            'slug': 'print_events_associated_with_invoices',
            'verbose_name': 'Invoice Check',
            'form_method': 'GET',
            'view': 'print_events_associated_with_invoices',
        },
    ]

    def merge_invoices(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request, "Vous n'êtes pas autorisé à effectuer cette action.",
                              level=messages.ERROR)
            return
        if queryset.count() < 2:
            self.message_user(request, "Vous devez sélectionner au moins deux lots de factures.",
                              level=messages.ERROR)
            return
        # merge all invoices in one
        # if one of the batch has a different type, we cannot merge
        for batch in queryset:
            if batch.batch_type != queryset[0].batch_type:
                self.message_user(request, "Vous ne pouvez pas fusionner des lots de factures de types différents.",
                                  level=messages.ERROR)
                return
        batch_ids = []
        youngest_start_date = queryset[0].start_date
        for invoice_batch in queryset:
            batch_ids.append(invoice_batch.id)
        new_invoice_batch = InvoiceItemBatch.objects.create(start_date=queryset[0].start_date,
                                                            end_date=queryset[0].end_date,
                                                            batch_type=queryset[0].batch_type,
                                                            batch_description="Merge of %s" % batch_ids)
        for batch in queryset:
            # attach all invoices to the new batch
            for invoice in batch.get_invoices():
                invoice.batch = new_invoice_batch
                invoice.save()

            # delete the batch
            # batch.delete()
        self.message_user(request, "Les lots de factures sélectionnés ont été fusionnés.",
                          level=messages.INFO)

    def print_events_associated_with_invoices(self, request, object_id, form_url='', extra_context=None, action=None):
        from django.template.response import TemplateResponse
        obj = self.get_object(request, object_id)
        return TemplateResponse(request, 'invoicing/print_events_associated_with_invoices.html', {'obj': obj})


@admin.register(InvoiceItemEmailLog)
class InvoiceItemEmailLogAdmin(admin.ModelAdmin):
    list_display = ('item', 'sent_at', 'recipient', 'subject')
    list_filter = ('sent_at',)
    search_fields = ('item__invoice_number', 'recipient', 'subject')
    ## display the foreign key as a link
    raw_id_fields = ('item',)
    ## all fields are read-only
    readonly_fields = [f.name for f in InvoiceItemEmailLog._meta.fields]


class TimesheetDetailInline(admin.TabularInline):
    extra = 1
    model = TimesheetDetail
    fields = ('start_date', 'end_date', 'task_description', 'patient',)
    search_fields = ['patient']
    ordering = ['start_date']


@admin.register(Timesheet)
class TimesheetAdmin(admin.ModelAdmin):
    fields = ('start_date', 'end_date', 'submitted_date', 'other_details', 'timesheet_validated')
    date_hierarchy = 'end_date'
    inlines = [TimesheetDetailInline]
    list_display = ('start_date', 'end_date', 'timesheet_owner', 'timesheet_validated')
    list_filter = ['employee', ]
    list_select_related = True
    readonly_fields = ('timesheet_validated',)
    verbose_name = 'Time sheet simple'
    verbose_name_plural = 'Time sheets simples'

    def save_model(self, request, obj, form, change):
        if not change:
            current_user = Employee.objects.raw(
                'select * from invoices_employee where user_id = %s' % (request.user.id))
            obj.employee = current_user[0]
        obj.save()

    def timesheet_owner(self, instance):
        return instance.employee.user.username

    def get_queryset(self, request):
        qs = super(TimesheetAdmin, self).get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.filter(employee__user=request.user)


class PublicHolidayCalendarDetailInline(admin.TabularInline):
    extra = 1
    model = PublicHolidayCalendarDetail


@admin.register(PublicHolidayCalendar)
class PublicHolidayCalendarAdmin(admin.ModelAdmin):
    inlines = [PublicHolidayCalendarDetailInline]
    verbose_name = u'Congés légaux'
    verbose_name_plural = u'Congés légaux'


class AbsenceRequestFileInline(admin.TabularInline):
    extra = 1
    model = AbsenceRequestFile


@admin.register(HolidayRequest)
class HolidayRequestAdmin(admin.ModelAdmin):
    class Media:
        css = {
            'all': ('css/holiday_request.css',)
        }

    date_hierarchy = 'start_date'
    list_filter = ('employee', FilteringYears, FilteringMonths, 'request_status', 'reason')
    ordering = ['-start_date']
    verbose_name = u"Demande d'absence"
    verbose_name_plural = u"Demandes d'absence"
    inlines = [AbsenceRequestFileInline]

    def holiday_request_status(self, obj):
        if HolidayRequestWorkflowStatus.ACCEPTED == obj.request_status:
            return format_html(
                '<div class="success">%s</div>' % HolidayRequestWorkflowStatus(obj.request_status).name)
        elif HolidayRequestWorkflowStatus.REFUSED == obj.request_status:
            return format_html(
                '<div class="error">%s</div>' % HolidayRequestWorkflowStatus(obj.request_status).name)
        else:
            return format_html(
                '<div class="warn">%s</div>' % HolidayRequestWorkflowStatus(obj.request_status).name)

    def sanity_check(self, obj):
        if obj.start_date.year != obj.end_date.year:
            return format_html(
                '<div class="warn">Attention, la demande concerne deux années différentes.</div>')
        else:
            return ''

    readonly_fields = ('validated_by', 'employee', 'request_creator', 'force_creation',
                       'request_status', 'validator_notes', 'total_days_in_current_year', 'do_not_notify',
                       'total_hours_off_available')
    list_display = ('employee', 'start_date', 'end_date', 'reason', 'hours_taken', 'validated_by',
                    'holiday_request_status', 'request_creator', 'total_days_in_current_year',
                    'sanity_check', 'total_hours_off_available')

    def accept_request(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request, "Vous n'avez pas le droit de valider des %s." % self.verbose_name_plural,
                              level=messages.WARNING)
            return
        rows_updated = 0
        obj: HolidayRequest
        for obj in queryset:
            if obj.request_status != HolidayRequestWorkflowStatus.ACCEPTED:
                try:
                    employee = Employee.objects.get(user_id=request.user.id)
                    obj.validated_by = employee
                    obj.request_status = HolidayRequestWorkflowStatus.ACCEPTED
                    if obj.validator_notes and len(obj.validator_notes) > 0:
                        obj.validator_notes = obj.validator_notes + "\n status: %s by %s on %s" % (
                            obj.request_status,
                            obj.validated_by,
                            timezone.localtime().strftime(
                                '%Y-%m-%dT%H:%M:%S'))
                    else:
                        obj.validator_notes = "status: %s by %s on %s" % (
                            obj.request_status,
                            obj.validated_by,
                            timezone.localtime().strftime(
                                '%Y-%m-%dT%H:%M:%S'))
                    obj.save()
                    notify_holiday_request_validation(obj, request)
                    rows_updated = rows_updated + 1
                except Employee.DoesNotExist:
                    self.message_user(request, "Vous n'avez de profil employé sur l'application pour valider une %s." %
                                      self.verbose_name_plural,
                                      level=messages.ERROR)
                    return
            else:
                ## if request accepted just pass
                pass
        if rows_updated == 1:
            message_bit = u"1 %s a été" % self.verbose_name
        else:
            message_bit = u"%s %s ont été" % (rows_updated, self.verbose_name_plural)
        self.message_user(request, u"%s validé avec succès." % message_bit)

    def refuse_request(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request, "Vous n'avez pas le droit de refuser des %s." % self.verbose_name_plural,
                              level=messages.WARNING)
            return
        rows_updated = 0
        obj: HolidayRequest
        for obj in queryset:
            if obj.request_status != HolidayRequestWorkflowStatus.REFUSED:
                try:
                    employee = Employee.objects.get(user_id=request.user.id)
                    obj.validated_by = employee
                    obj.request_status = HolidayRequestWorkflowStatus.REFUSED
                    obj.request_creator = request.user
                    if len(obj.validator_notes) > 0:
                        obj.validator_notes = obj.validator_notes + "\n status: %s by %s on %s" % (
                            obj.request_status,
                            obj.validated_by,
                            timezone.localtime().strftime(
                                '%Y-%m-%dT%H:%M:%S'))
                    else:
                        obj.validator_notes = "status: %s by %s on %s" % (
                            obj.request_status,
                            obj.validated_by,
                            timezone.localtime().strftime(
                                '%Y-%m-%dT%H:%M:%S'))
                    if not obj.request_creator:
                        obj.request_creator = obj.employee
                    obj.save()
                    notify_holiday_request_validation(obj, request)
                    rows_updated = rows_updated + 1
                except Employee.DoesNotExist:
                    self.message_user(request,
                                      "Vous n'avez de profil employé sur l'application pour refuser une %s." %
                                      self.verbose_name_plural,
                                      level=messages.ERROR)
                    return
            else:
                ## if request already refused just pass
                pass
        if rows_updated == 1:
            message_bit = u"1 %s a été" % self.verbose_name
        else:
            message_bit = u"%s %s ont été" % (rows_updated, self.verbose_name_plural)
        self.message_user(request, u"%s refusée(s) avec succès." % message_bit)

    def response_change(self, request, obj):
        queryset = HolidayRequest.objects.filter(id=obj.id)
        if "_accept_request" in request.POST:
            self.accept_request(request, queryset)
            return HttpResponseRedirect(request.path)
        if "_refuse_request" in request.POST:
            self.refuse_request(request, queryset)
            return HttpResponseRedirect(request.path)
        return HttpResponseRedirect(request.path)

    def get_readonly_fields(self, request, obj=None):
        if obj is not None:
            if obj.employee.employee.user.id != request.user.id and not 1 != obj.request_status and not request.user.is_superuser:
                return 'employee', 'start_date', 'end_date', 'requested_period', 'reason', 'validated_by', \
                    'hours_taken', 'request_creator'
            elif request.user.is_superuser and not 1 != obj.request_status:
                return [f for f in self.readonly_fields if f not in ['force_creation', 'do_not_notify', 'employee',
                                                                     'request_status',
                                                                     'validator_notes']]
            elif request.user.is_superuser:
                return [f for f in self.readonly_fields if f not in ['force_creation', 'do_not_notify', 'employee',
                                                                     'validator_notes']]
            elif 0 != obj.request_status:
                return 'employee', 'start_date', 'end_date', 'requested_period', 'reason', 'validated_by', \
                    'hours_taken', 'request_creator', 'request_status', 'validator_notes', 'force_creation', 'do_not_notify'

        else:
            if request.user.is_superuser:
                return [f for f in self.readonly_fields if f not in ['force_creation', 'do_not_notify', 'employee',
                                                                     'request_status', 'validator_notes']]
        return self.readonly_fields

    def has_delete_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True
        if obj:
            return request.user.is_superuser or (obj.employee.id == request.user.id and not 1 != obj.request_status)
        else:
            if 'object_id' in request.resolver_match:
                object_id = request.resolver_match.kwargs['object_id']
                holiday_request = HolidayRequest.objects.get(id=object_id)
                return request.user.is_superuser or (
                        holiday_request.employee.id == request.user.id and not 1 != holiday_request.request_status)

    def get_actions(self, request):
        actions = super().get_actions(request)
        if not request.user.is_superuser:
            if 'validate_or_invalidate_request' in actions:
                del actions['validate_or_invalidate_request']
        return actions


class SimplifiedTimesheetDetailInline(admin.TabularInline):
    extra = 1
    model = SimplifiedTimesheetDetail
    # fields = ('start_date', 'end_date', 'time_delta')
    readonly_fields = ('time_delta',)
    # 'coordinates_at_start_date', 'coordinates_at_end_date')
    ordering = ['start_date']
    formset = SimplifiedTimesheetDetailForm
    # template = 'admin/invoices/simplifiedtimesheetdetail/tabular.html'
    # class Media:
    #     js = [
    #         'js/tsheet_update_location.js',
    #     ]


def calculate_balance_for_previous_months(month, year, user_id):
    today = datetime.date.today()
    first = today.replace(month=month, year=year).replace(day=1)
    lastMonth = first - datetime.timedelta(days=1)
    previous_timsheets = SimplifiedTimesheet.objects.filter(time_sheet_month=lastMonth.month,
                                                            time_sheet_year=lastMonth.year,
                                                            employee__user_id=user_id,
                                                            timesheet_validated=True)
    for prev_months_tsheet in previous_timsheets:
        return Decimal(round(prev_months_tsheet.hours_should_work_gross_in_sec / 3600, 2)) \
            - prev_months_tsheet.extra_hours_paid_current_month \
            + prev_months_tsheet.extra_hours_balance
    return 0


@admin.register(SimplifiedTimesheet)
class SimplifiedTimesheetAdmin(CSVExportAdmin):
    ordering = ('-time_sheet_year', '-time_sheet_month')
    inlines = [SimplifiedTimesheetDetailInline]
    csv_fields = ['employee', 'time_sheet_year', 'time_sheet_month',
                  'simplifiedtimesheetdetail__start_date',
                  'simplifiedtimesheetdetail__end_date']
    list_display = ('timesheet_owner', 'timesheet_validated', 'time_sheet_year', 'time_sheet_month',
                    'extra_hours_balance')
    list_filter = ['employee', 'time_sheet_year', 'time_sheet_month']
    list_select_related = True
    readonly_fields = ('timesheet_validated', 'total_hours',
                       'total_hours_sundays', 'total_hours_public_holidays', 'total_working_days',
                       'total_legal_working_hours',
                       'total_hours_holidays_taken', 'hours_should_work', 'extra_hours_balance',
                       'extra_hours_paid_current_month', 'created_on', 'updated_on')
    verbose_name = 'Temps de travail'
    verbose_name_plural = 'Temps de travail'
    actions = ['validate_time_sheets', 'timesheet_situation', 'force_cache_clearing', 'build_use_case_objects']
    form = SimplifiedTimesheetForm

    def build_use_case_objects(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request,
                              "Vous n'avez pas le droit de faire cette action des %s." % self.verbose_name_plural)
            return
        file_data = build_use_case_objects(queryset)
        response = HttpResponse(file_data, content_type='application/text charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="test_case.txt"'
        return response

    def timesheet_situation(self, request, queryset):
        _counter = 1
        file_data = ""
        for tsheet in queryset:
            if not tsheet.timesheet_validated:
                self.message_user(request,
                                  "Timesheet %s n'est pas validée, vous devez la valider avant de pouvoir exporter le "
                                  "rapport" % tsheet,
                                  level=messages.ERROR)
                return
            else:
                # take previous months timesheet
                first = tsheet.get_start_date
                lastMonth = first - datetime.timedelta(days=1)
                previous_timsheets = SimplifiedTimesheet.objects.filter(time_sheet_month=lastMonth.month,
                                                                        time_sheet_year=lastMonth.year,
                                                                        employee__user_id=tsheet.user.id,
                                                                        timesheet_validated=True)
                if len(previous_timsheets) == 0:
                    file_data += "\n {counter} - {last_name} {first_name}:\n".format(counter=_counter,
                                                                                     last_name=tsheet.user.last_name.upper(),
                                                                                     first_name=tsheet.user.first_name)
                    if tsheet.extra_hours_paid_current_month:
                        file_data += "\nA travaillé {total_extra} heures supplémentaires.".format(
                            total_extra=tsheet.extra_hours_paid_current_month)
                    if tsheet.total_hours_holidays_and_sickness_taken > 0:
                        file_data += "\n{holidays_sickness_explanation}".format(
                            holidays_sickness_explanation=tsheet.total_hours_holidays_and_sickness_taken_object.beautiful_explanation())
                else:
                    previous_month_tsheet = previous_timsheets.first()
                    file_data += "\n {counter} - {last_name} {first_name}:\n".format(counter=_counter,
                                                                                     last_name=tsheet.user.last_name.upper(),
                                                                                     first_name=tsheet.user.first_name)
                    if tsheet.total_hours_sundays:
                        file_data += " \nA travaillé {hours_sunday} heures des Dimanche".format(
                            hours_sunday=previous_month_tsheet.total_hours_sundays)
                    if tsheet.total_hours_public_holidays:
                        file_data += " \nA travaillé {total_hours_public_holidays} heures des Jours fériés.".format(
                            total_hours_public_holidays=previous_month_tsheet.total_hours_public_holidays)
                    if tsheet.extra_hours_paid_current_month:
                        file_data += " \nA travaillé {total_extra} heures supplémentaires.".format(
                            total_extra=tsheet.extra_hours_paid_current_month)
                    if tsheet.absence_hours_taken()[0] > 0:
                        holiday_sickness_explanation = tsheet.absence_hours_taken()[1].beautiful_explanation()
                        if len(holiday_sickness_explanation) > 0:
                            file_data += holiday_sickness_explanation
                    if previous_month_tsheet.absence_hours_taken()[0] > 0:
                        file_data += "\nPour rappel le mois de %s" % previous_month_tsheet.get_start_date.month
                        holiday_sickness_explanation = previous_month_tsheet.absence_hours_taken()[
                            1].beautiful_explanation()
                        if len(holiday_sickness_explanation) > 0:
                            file_data += holiday_sickness_explanation

            _counter += 1
        response = HttpResponse(file_data, content_type='application/text charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="timesheet_situation.txt"'
        return response

    def get_readonly_fields(self, request, obj=None):
        if obj is not None:
            if obj.timesheet_validated:
                return self.readonly_fields
            elif request.user.is_superuser and not obj.timesheet_validated:
                return tuple(x for x in self.readonly_fields if x not in ('extra_hours_paid_current_month',
                                                                          'extra_hours_balance'))
        return self.readonly_fields

    def save_model(self, request, obj, form, change):
        if not change:
            obj.employee = Employee.objects.get(user_id=obj.user.id)
        obj.save()

    def timesheet_owner(self, instance):
        return instance.employee.user.username

    def get_queryset(self, request):
        qs = super(SimplifiedTimesheetAdmin, self).get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.filter(employee__user=request.user)

    def validate_time_sheets(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request, "Vous n'avez pas le droit de valider des %s." % self.verbose_name_plural,
                              level=messages.WARNING)
            return
        rows_updated = 0
        obj: SimplifiedTimesheet
        for obj in queryset:
            obj.timesheet_validated = not obj.timesheet_validated
            rows_updated = rows_updated + 1
            obj.extra_hours_balance = calculate_balance_for_previous_months(month=obj.time_sheet_month,
                                                                            year=obj.time_sheet_year,
                                                                            user_id=obj.employee.user.id)
            obj.save()

        if rows_updated == 1:
            message_bit = u"1 time sheet a été"
        else:
            message_bit = u"%s time sheet ont été" % rows_updated
        self.message_user(request, u"%s (in)validé avec succès." % message_bit)

    def change_view(self, request, object_id, extra_context=None):
        if SimplifiedTimesheet.objects.get(pk=object_id).timesheet_validated and not request.user.is_superuser:
            extra_context = extra_context or {}
            extra_context['readonly'] = True
        return super(SimplifiedTimesheetAdmin, self).change_view(request, object_id, extra_context=extra_context)

    def has_change_permission(self, request, obj=None):
        if obj and obj.timesheet_validated and not request.user.is_superuser:
            return False
        return self.has_change_permission

    def has_delete_permission(self, request, obj=None):
        if obj and obj.timesheet_validated and not request.user.is_superuser:
            return False
        return self.has_delete_permission

    def force_cache_clearing(self, request, queryset):
        if request.user.is_superuser:
            cache.clear()
            self.message_user(request, u"Cache refresh OK.", level=messages.INFO)
        else:
            self.message_user(request, u"Not super user.", level=messages.WARNING)


@admin.register(EventType)
class EventTypeAdmin(admin.ModelAdmin):
    list_display = ['name']


class AssignedAdditionalEmployeeInLine(admin.StackedInline):
    extra = 0
    model = AssignedAdditionalEmployee
    fields = ('assigned_additional_employee',)
    autocomplete_fields = ['assigned_additional_employee']


class ReportPictureInLine(admin.StackedInline):
    extra = 0
    model = ReportPicture
    fields = ('description', 'image',)

    def has_add_permission(self, request, obj):
        return True

    def has_change_permission(self, request, obj=None):
        return True

    def has_delete_permission(self, request, obj=None):
        return True

    def has_view_permission(self, request, obj=None):
        return True


class GenericTaskDescriptionInline(admin.TabularInline):
    extra = 0
    model = GenericTaskDescription


class EventLinkToCareCodeInline(admin.TabularInline):
    extra = 0
    model = EventLinkToCareCode
    autocomplete_fields = ['care_code']


class EventLinkToMedicalCareSummaclryPerPatientDetailInline(admin.TabularInline):
    extra = 0
    model = EventLinkToMedicalCareSummaryPerPatientDetail
    formset = EventLinkToMedicalCareSummaryPerPatientDetailForm

    def get_queryset(self, request):

        qs = super().get_queryset(request)
        if 'object_id' in request.resolver_match.kwargs:
            event_instance = self.parent_model.objects.get(id=request.resolver_match.kwargs['object_id'])
            linked_patient = event_instance.patient
            #for instance in qs:
            #    print(
            #        f'MedicalCareSummaryPerPatientDetail patient: {instance.medical_care_summary_per_patient_detail.medical_care_summary_per_patient.patient}')
            # from dependence.detailedcareplan import MedicalCareSummaryPerPatientDetail
            # return MedicalCareSummaryPerPatientDetail.objects.filter(medical_care_summary_per_patient__patient=linked_patient)
            return qs.filter(
                medical_care_summary_per_patient_detail__medical_care_summary_per_patient__patient=linked_patient)
        else:
            return qs.none()  # Return an empty queryset

    # hide field quantity to non superuser
    def get_readonly_fields(self, request, obj=None):
        if not request.user.is_superuser:
            return ['medical_care_summary_per_patient_detail', 'quantity']
        return self.readonly_fields


class EventGenericLinkInline(admin.TabularInline):
    model = EventGenericLink
    extra = 0  # Number of extra forms to display


@admin.register(Event)
class EventAdmin(admin.ModelAdmin):
    class Media:
        css = {
            "all": ("css/event.css",)
        }
        js = [
            "js/conditional-event-address.js",
            "js/toggle-period.js",
        ]

    form = EventForm

    list_display = ['day', 'state', 'patient', 'event_type_enum', 'notes', ]
    readonly_fields = ['created_by', 'created_on', 'calendar_url', 'calendar_id']
    exclude = ('event_type',)
    autocomplete_fields = ['patient']
    change_list_template = 'events/change_list.html'
    change_form_template = 'admin/invoices/event_change_form.html'
    list_filter = (SmartEmployeeFilter,)
    inlines = (ReportPictureInLine, EventLinkToCareCodeInline,
               EventLinkToMedicalCareSummaclryPerPatientDetailInline,
               EventGenericLinkInline,
               GenericTaskDescriptionInline)

    def get_form(self, request, obj=None, change=False, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if not request.user.is_superuser and not request.user.groups.filter(name="planning manager").exists():
            if len(form.base_fields) > 0:
                form.base_fields["event_report"].required = True
                form.base_fields["state"].choices = (3, _('Done')), (5, _('Not Done'))

        class ModelFormWithRequest(form):
            def __new__(cls, *args, **kwargs):
                kwargs['request'] = request
                return form(*args, **kwargs)

        return ModelFormWithRequest

    def get_readonly_fields(self, request, obj=None):
        if obj is not None:
            if not request.user.is_superuser and not request.user.groups.filter(name="planning manager").exists():
                fs = [field.name for field in Event._meta.fields if field.name != "id"]
                if obj.employees is not None and obj.employees.user.id == request.user.id:
                    return [f for f in fs if f not in ['event_report', 'state']]
                else:
                    return fs
        return self.readonly_fields

    @csrf_protect_m
    def changelist_view(self, request, extra_context=None):
        after_day = request.GET.get('day__gte', None)
        extra_context = extra_context or {}
        employee_id = request.GET.get('employee', None)
        # period = request.GET.get('period', None)

        if not after_day:
            d = datetime.date.today()
        else:
            try:
                split_after_day = after_day.split('-')
                d = datetime.date(year=int(split_after_day[0]), month=int(split_after_day[1]), day=1)
            except:
                d = datetime.date.today()

        previous_month = datetime.date(year=d.year, month=d.month, day=1)  # find first day of current month
        previous_month = previous_month - datetime.timedelta(days=1)  # backs up a single day
        previous_month = datetime.date(year=previous_month.year, month=previous_month.month,
                                       day=1)  # find first day of previous month

        last_day = calendar.monthrange(d.year, d.month)
        next_month = datetime.date(year=d.year, month=d.month, day=last_day[1])  # find last day of current month
        next_month = next_month + datetime.timedelta(days=1)  # forward a single day
        next_month = datetime.date(year=next_month.year, month=next_month.month,
                                   day=1)  # find first day of next month

        extra_context['list_view'] = reverse('admin:invoices_eventlist_changelist')
        extra_context['previous_month'] = reverse('admin:invoices_event_changelist') + '?day__gte=' + str(
            previous_month)
        extra_context['next_month'] = reverse('admin:invoices_event_changelist') + '?day__gte=' + str(next_month)

        cal: EventCalendar = EventCalendar()
        # if "month" == period:
        html_calendar = cal.formatmonth(d.year, d.month, withyear=True, employee_id=employee_id)
        # else:
        #     # html_calendar = cal.formatweek(d.year, d.month, withyear=True, employee_id=employee_id)
        #     html_calendar = cal.formatmonth(d.year, d.month, withyear=True, employee_id=employee_id)
        html_calendar = html_calendar.replace('<td ', '<td  width="150" height="150"')
        extra_context['calendar'] = mark_safe(html_calendar)
        return super(EventAdmin, self).changelist_view(request, extra_context)


@admin.register(EventList)
class EventListAdmin(admin.ModelAdmin):
    list_display = ['day', 'time_start_event', 'time_end_event', 'patient', 'employees', 'state',
                    'created_on', 'aevs_stats']
    change_list_template = 'admin/change_list.html'
    list_filter = (
        'employees', 'event_type_enum', 'state', SmartPatientFilter, 'created_by', UnderAssuranceDependanceFilter)
    date_hierarchy = 'day'
    date_hierarchy_format = '%Y-%m-%d'
    exclude = ('event_type',)

    actions = ['safe_delete', 'duplicate_event_for_next_day', 'duplicate_event_for_next_week',
               "duplicate_event_for_the_hole_week",
               'delete_in_google_calendar', 'list_orphan_events', 'force_gcalendar_sync',
               'cleanup_events_event_types', 'print_unsynced_events', 'cleanup_all_events_on_google',
               'send_webhook_message', 'export_to_excel']

    inlines = (ReportPictureInLine, EventLinkToCareCodeInline,
               EventLinkToMedicalCareSummaclryPerPatientDetailInline,
               EventGenericLinkInline,
               GenericTaskDescriptionInline)
    search_fields = ['patient__first_name', 'patient__name', 'patient__phone_number', 'patient__email_address',
                     'notes', 'event_report', 'calendar_id']
    autocomplete_fields = ['patient']

    form = EventForm

    def export_to_excel(self, request, queryset):
        # Check if the user is a superuser
        if not request.user.is_superuser:
            self.message_user(request, "Vous n'êtes pas autorisé à effectuer cette action.", level=messages.ERROR)
            return

        # Check that all events are for the same patient
        if len(set([e.patient for e in queryset])) > 1:
            self.message_user(request, "Tous les événements doivent être pour le même patient.", level=messages.ERROR)
            return

        # Check that all events are for the same year/month
        if len(set([e.day.year for e in queryset])) > 1 or len(set([e.day.month for e in queryset])) > 1:
            self.message_user(request, "Tous les événements doivent être pour le même mois.", level=messages.ERROR)
            return

        # Create the Excel response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="validation.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active

        # Set the sheet title to month and year
        ws.title = queryset.first().day.strftime('%m-%Y')

        # Determine the month and year from the queryset
        if queryset.exists():
            event_month = queryset.first().day.month
            event_year = queryset.first().day.year
        else:
            self.message_user(request, "No events selected.", level=messages.ERROR)
            return

        # Define fill patterns for alternating colors
        fill_color_1 = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        fill_color_2 = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")

        # Initialize a flag to alternate colors
        use_color_1 = True

        # Create a bold and larger font style
        bold_larger_font = Font(size=12, bold=True)
        # Create a bold font style
        bold_font = Font(bold=True)

        # Append a row with the name of the patient in bold and the matricule
        patient_cell = ws.cell(row=ws.max_row + 1, column=1, value="Patient")
        patient_cell.font = bold_larger_font
        patient_name_cell = ws.cell(row=ws.max_row, column=2, value=str(queryset.first().patient))
        patient_name_cell.font = bold_larger_font

        matricule_cell = ws.cell(row=ws.max_row + 1, column=1, value="Matricule")
        matricule_cell.font = bold_larger_font
        matricule_value_cell = ws.cell(row=ws.max_row, column=2, value=queryset.first().patient.code_sn)
        matricule_value_cell.font = bold_larger_font

        # Add an empty row
        ws.append([])

        # Append a row with validations of the month of the events selected in bold styling
        validation_month_cell = ws.cell(row=ws.max_row + 2, column=1, value="Validation du mois")
        validation_month_cell.font = bold_larger_font
        validation_month_value_cell = ws.cell(row=ws.max_row, column=2, value=queryset.first().day.strftime('%m %Y'))
        validation_month_value_cell.font = bold_larger_font
        ws.append([])

        # Prepare headers with dates of the month
        days_in_month = calendar.monthrange(event_year, event_month)[1]
        headers = ['Care Code']
        for day in range(1, days_in_month + 1):
            headers.append(datetime.date(event_year, event_month, day).strftime('%d'))
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = bold_font

        # Aggregate data
        care_code_data = defaultdict(lambda: defaultdict(int))  # {care_code: {day: quantity}}
        for event in queryset:
            for link in event.eventlinktomedicalcaresummaryperpatientdetail_set.all():
                care_code_data[link.medical_care_summary_per_patient_detail.item.code][event.day] += link.quantity

        # Sort and group codes by their starting letters
        grouped_codes = defaultdict(list)
        for care_code in care_code_data.keys():
            starting_letter = care_code[0]
            grouped_codes[starting_letter].append(care_code)

        # Populate the worksheet with grouped codes
        for starting_letter in sorted(grouped_codes.keys()):
            ws.append([])  # Add an empty row before each new group
            for care_code in sorted(grouped_codes[starting_letter]):
                day_data = care_code_data[care_code]
                row = [care_code] + [" "] * days_in_month  # Initialize row with zeros for each day
                for day, quantity in day_data.items():
                    day_index = day.day - 1
                    row[day_index + 1] = quantity
                ws.append(row)
                current_row = ws.max_row
                for cell in ws[current_row]:
                    # Apply alternating color fills
                    cell.fill = fill_color_1 if use_color_1 else fill_color_2
                # Toggle the flag for the next row
                use_color_1 = not use_color_1

        column_width = 20  # Set the desired width for the first column
        ws.column_dimensions['A'].width = column_width

        wb.save(response)
        return response


    @transaction.atomic
    def create_invoice_item_out_of_events(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request, "Vous n'avez pas le droit de créer des factures.", level=messages.WARNING)
            return
        # All events must be for the same patient
        if len(set([e.patient for e in queryset])) > 1:
            self.message_user(request, "Tous les événements doivent être pour le même patient.", level=messages.WARNING)
            return
        # get the patient
        patient = queryset[0].patient
        prestations = []
        default_invoicing_dtls = InvoicingDetails.objects.get(default_invoicing=True)
        date_31_01_2024 = datetime.date(2024, 1, 31)
        # one invoice can only hold 20 prestation, so we need to  one invoice per 20 event bunch
        if len(queryset) > 10:
            invoices_created = []
            # create a new invoice item for each 20 events
            for i in range(0, len(queryset), 10):
                invoice_item = InvoiceItem.objects.create(
                    invoice_date=date_31_01_2024,
                    invoice_details=default_invoicing_dtls,
                    patient=patient)
                invoices_created.append(invoice_item)
                for event in queryset[i:i + 10]:
                    prestations.append(event.create_prestation_out_of_event(invoice_item=invoice_item, ))
            self.message_user(request, "Factures %s créées pour %s" % (invoices_created, patient))
        else:
            invoice_item = InvoiceItem.objects.create(
                invoice_date=date_31_01_2024,
                invoice_details=default_invoicing_dtls,
                patient=patient)
            for event in queryset:
                event.create_prestation_out_of_event(invoice_item=invoice_item)
            self.message_user(request, "Facture %s créée pour %s" % (invoice_item, patient))

    # create an action that will filter events that have no prestation associated to them on date
    def filter_events_with_no_invoiced_prestation_on_date(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request,
                              "Vous n'avez pas le droit de lister les %s orphelins." % self.verbose_name_plural,
                              level=messages.WARNING)
            return

        events = Event.objects.filter(day__range=(self.start_date, self.end_date)).exclude(patient__isnull=True).filter(
            patient__is_under_dependence_insurance=False).exclude(event_type_enum=EventTypeEnum.BIRTHDAY).order_by(
            "patient__name", "day")
        # grouped_events = events.values('patient__name').annotate(event_count=Count('id'))
        for e in queryset:
            e.display_unconnected_events()

    def safe_delete(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request, "Vous n'avez pas le droit de supprimer des %s." % self.verbose_name_plural,
                              level=messages.WARNING)
            return
        for e in queryset:
            e.delete()

    def send_webhook_message(self, request, queryset):
        if not request.user.is_superuser:
            return
        for e in queryset:
            post_webhook(employees=e.employees, patient=e.patient, event_report=e.event_report, state=e.state,
                         event_date=datetime.datetime.combine(e.day, e.time_start_event).astimezone(ZoneInfo("Europe"
                                                                                                             "/Luxembourg")))

    def list_orphan_events(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request,
                              "Vous n'avez pas le droit de lister les %s orphelins." % self.verbose_name_plural,
                              level=messages.WARNING)
            return
        for e in queryset:
            if isinstance(e, EventList):
                time_min = datetime.datetime.combine(e.day, datetime.time(0, 0, 0)).astimezone(ZoneInfo("Europe"
                                                                                                        "/Luxembourg"))
                if os.environ.get('LOCAL_ENV', None):
                    e.display_unconnected_events(time_min)
                else:
                    e.display_unconnected_events.delay(time_min)

    def duplicate_event_for_next_week(self, request, queryset):
        if not request.user.is_superuser and not request.user.groups.filter(name="planning manager").exists():
            self.message_user(request, "Vous n'avez pas le droit de dupliquer des %s." % self.model._meta.verbose_name_plural,
                              level=messages.WARNING)

            return
        # only one event at a time
        events_duplicated = []
        if len(queryset) < 6:
            print("duplicating %s events [direct call]" % len(queryset))
            for e in queryset:
                events_duplicated.append(e.duplicate_event_for_next_day(7))
            self.message_user(request, "Duplicated %s events" % len(events_duplicated))
            # redirect to list filtering on duplicated events
            return HttpResponseRedirect(
                reverse('admin:invoices_eventlist_changelist') + '?id__in=' + ','.join(
                    [str(e.id) for e in events_duplicated]))
        else:
            print("duplicating %s events [redis rq call]" % len(queryset))
            # create array with all selected events
            from invoices.processors.tasks import duplicate_event_for_next_day_for_several_events
            duplicate_event_for_next_day_for_several_events.delay([e for e in queryset], request.user, number_of_days=7)
            self.message_user(request,
                              "Il y a %s événements à dupliquer pour j + 7, cela peut prendre quelques minutes, vous allez recevoir une notification par google chat à la fin de la création" % len(
                                  queryset))

    @transaction.atomic
    def create_assurance_dependance_invoice_out_of_events(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request, "Vous n'avez pas le droit de créer des factures d'assurance dépendance.",
                              level=messages.WARNING)
            return
        if len(set([e.patient for e in queryset])) > 1:
            self.message_user(request, "Tous les événements doivent être pour le même patient. or vous avez sélectionné des patients suivants %s" % set([e.patient for e in queryset]), level=messages.WARNING)
            return
        # check that all events are in the same month
        if len(set([e.day.month for e in queryset])) > 1:
            self.message_user(request, "Tous les événements doivent être pour le même mois.", level=messages.WARNING)
            return
        # check event start is the first day of the month
        invoice_start_period = queryset[0].day.replace(day=1)
        # invoice end date should be the last day of the month
        invoice_end_period = queryset[0].day.replace(
            day=calendar.monthrange(queryset[0].day.year, queryset[0].day.month)[1])

        long_term_invoice = LongTermCareInvoiceFile.objects.create(invoice_start_period=invoice_start_period,
                                                                   invoice_end_period=invoice_end_period,
                                                                   patient=queryset[0].patient)
        # FIXME: hardcoded values
        long_term_care_package_amd_gi = LongTermPackage.objects.get(code="AMDGI")
        subcontractor = SubContractor.objects.get(name="OP DER SCHOCK a.s.b.l. et s.c.")
        for event in queryset:
            if event.state == 3:
                LongTermCareInvoiceItem.objects.create(invoice=long_term_invoice, care_date=event.day,
                                                       long_term_care_package=long_term_care_package_amd_gi,
                                                       quantity=event.duration_in_hours() * 2,
                                                       subcontractor=subcontractor)

            # long_term_invoice.add_event(event)
        long_term_invoice.save()
        self.message_user(request, "Invoice created for %s" % long_term_invoice)

    @transaction.atomic
    def duplicate_event_for_the_hole_week(self, request, queryset):
        if not request.user.is_superuser or not request.user.groups.filter(name="planning manager").exists():
            self.message_user(request, "Vous n'avez pas le droit de dupliquer des %s." % self.verbose_name_plural,
                              level=messages.WARNING)
            return
        events_duplicated = []
        if len(queryset) < 6:
            print("duplicating %s events [direct call]" % len(queryset))
            for e in queryset:
                events_duplicated = e.repeat_event_for_all_days_of_week(e.day)
            self.message_user(request, "Duplicated %s events" % len(events_duplicated))
            # redirect to list filtering on duplicated events
            return HttpResponseRedirect(
                reverse('admin:invoices_eventlist_changelist') + '?id__in=' + ','.join(
                    [str(e.id) for e in events_duplicated]))

    def duplicate_event_for_next_day(self, request, queryset):
        if not request.user.is_superuser and not request.user.groups.filter(name="planning manager").exists():
            self.message_user(request, "Vous n'avez pas le droit de dupliquer des %s." % self.model._meta.verbose_name_plural,
                              level=messages.WARNING)
            return
        # only one event at a time
        events_duplicated = []
        if len(queryset) < 6:
            print("duplicating %s events [direct call]" % len(queryset))
            for e in queryset:
                events_duplicated.append(e.duplicate_event_for_next_day(1))
            self.message_user(request, "Duplicated %s events" % len(events_duplicated))
            # redirect to list filtering on duplicated events
            return HttpResponseRedirect(
                reverse('admin:invoices_eventlist_changelist') + '?id__in=' + ','.join(
                    [str(e.id) for e in events_duplicated]))
        else:
            print("duplicating %s events [redis rq call]" % len(queryset))
            # create array with all selected events
            from invoices.processors.tasks import duplicate_event_for_next_day_for_several_events
            duplicate_event_for_next_day_for_several_events.delay([e for e in queryset], request.user, number_of_days=1)
            self.message_user(request,
                              "Il y a %s événements à dupliquer pour j + 1, cela peut prendre quelques minutes, vous allez recevoir une notification par google chat à la fin de la création" % len(
                                  queryset))

    def cleanup_all_events_on_google(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request, "Must be super user", level=messages.ERROR)
            return
        for e in queryset:
            result = e.cleanup_all_events_on_google(dry_run=False)
            self.message_user(request, "Deleted %s messages from Google calendar " % len(result),
                              level=messages.WARNING)

    def force_gcalendar_sync(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request, "Must be super user", level=messages.ERROR)
            return
        evts_synced = []
        for e in queryset:
            cal = create_or_update_google_calendar(e)
            e.calendar_id = cal.get('id')
            e.calendar_url = cal.get('htmlLink')
            evts_synced.append(e.calendar_url)
            e.save()
        self.message_user(request, "%s évenements synchronisés. : %s" % (len(evts_synced), evts_synced),
                          level=messages.INFO)

    def print_unsynced_events(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request, "Must be super user", level=messages.ERROR)
            return
        evts_not_synced = []
        for e in queryset:
            if 'http://a.sur.lu' == e.calendar_url and e.calendar_id == 0:
                evts_not_synced.append(e)
        if len(evts_not_synced) > 0:
            self.message_user(request, "%s évenements non synchronisés. : %s" % (len(evts_not_synced), evts_not_synced),
                              level=messages.WARNING)
        else:
            self.message_user(request, "Tous les évenements sont synchronisés.",
                              level=messages.INFO)

    def cleanup_events_event_types(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request, "Must be super user", level=messages.ERROR)
            return
        evts_cleaned = []
        for e in Event.objects.filter(event_type=EventType.objects.get(id=1)):
            if e.event_type_enum != EventTypeEnum.BIRTHDAY:
                e.event_type_enum = EventTypeEnum.BIRTHDAY
                e.save()
                evts_cleaned.append(e)
        if len(evts_cleaned) > 0:
            self.message_user(request, "%s événements nettoyés. : %s" % (len(evts_cleaned), evts_cleaned),
                              level=messages.WARNING)
        else:
            self.message_user(request, "Tous les évenements sont propres.",
                              level=messages.INFO)

    def get_queryset(self, request):
        queryset = super(EventListAdmin, self).get_queryset(request)
        today = dt.now()
        # Check if any filters have been applied
        if request.GET:
            return queryset
        if request.user.is_superuser or request.user.groups.filter(name="planning manager").exists():
            return Event.objects.all()
        else:
            # filter only events assigned to the current user and of today, can be also of yesterday
            filtered_events = queryset.filter(day=today, employees__user=request.user, state=2) | queryset.filter(
                day=today - datetime.timedelta(days=1), employees__user=request.user, state=2)
            return filtered_events

    def get_form(self, request, obj=None, change=False, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if not request.user.is_superuser or not request.user.groups.filter(name="planning manager").exists():
            if len(form.base_fields) > 0:
                form.base_fields["event_report"].required = True
                form.base_fields["state"].choices = (3, _('Done')), (5, _('Not Done')), (6, _('Cancelled'))
                form.base_fields.pop("calendar_url", None)
                form.base_fields.pop("calendar_id", None)
                form.base_fields.pop("created_by", None)
                form.base_fields.pop("created_on", None)
                form.base_fields.pop("updated_on", None)
                form.base_fields.pop("event_type", None)
        return form

        class ModelFormWithRequest(form):
            def __new__(cls, *args, **kwargs):
                kwargs['request'] = request
                return form(*args, **kwargs)

        return ModelFormWithRequest

    def get_readonly_fields(self, request, obj=None):
        if obj is not None:
            if not request.user.is_superuser:
                fs = [field.name for field in Event._meta.fields if field.name not in ("id")]
                if obj.employees and obj.employees.user.id == request.user.id:
                    return [f for f in fs if f not in ['event_report', 'state', 'calendar_id', "calendar_url",
                                                       "created_by",
                                                       "created_on", "event_type"]]
                elif obj.sub_contractor and obj.sub_contractor.name in [group.name for group in
                                                                        request.user.groups.all()]:
                    return [f for f in fs if
                            f not in ['event_report', 'state', 'calendar_id', "updated_on", "calendar_url",
                                      "created_by", "created_on", "event_type"]]
                else:
                    return fs
        return self.readonly_fields

    def changelist_view(self, request, extra_context=None):
        changelist = ChangeList(request,
                                self.model,
                                list(self.get_list_display(request)),
                                self.get_list_display_links(request, self.get_list_display(request)),
                                self.get_list_filter(request),
                                self.date_hierarchy,
                                self.get_search_fields(request),
                                self.list_select_related,
                                self.list_per_page,
                                self.list_max_show_all,
                                self.list_editable,
                                self,
                                self.sortable_by,
                                self.get_search_results)

        # Get a distinct list of patients for the filtered queryset
        patients = list(set(changelist.get_queryset(request).values_list('patient__id', 'patient__name')))

        # Attach it to request
        request.dynamic_patient_choices = [(str(patient_id), patient_name) for patient_id, patient_name in patients]

        return super().changelist_view(request, extra_context)


class EventWeekList(Event):
    class Meta:
        proxy = True
        verbose_name = "Nouveau Planning *"
        verbose_name_plural = "Nouveaux Plannings *"


@admin.register(EventWeekList)
class EventWeekListAdmin(admin.ModelAdmin):
    class Media:
        css = {
            "all": ("css/main.min.css",
                    "css/fullcalendar-additional.css")
        }
        js = [
            "js/main.min.js",
        ]

    list_display = ['day', 'time_start_event', 'time_end_event', 'state', 'event_type_enum', 'patient', 'employees']
    exclude = ('event_type',)
    change_list_template = 'events/calendar.html'
    change_form_template = 'admin/invoices/event_change_form.html'
    list_filter = ('employees', 'event_type_enum', 'state', 'patient', 'created_by')
    date_hierarchy = 'day'

    actions = ['safe_delete', 'delete_in_google_calendar', 'list_orphan_events', 'create_invoice_item_out_of_events', ]
    readonly_fields = ['created_by', 'created_on', 'updated_on', 'calendar_url', 'calendar_id']

    context = {}
    form = EventForm

    @csrf_protect_m
    def changelist_view(self, request, extra_context=None):
        raw_list = Event.objects.filter(day__month=datetime.date.today().month).order_by("employees_id")
        # object_list = (Event.objects
        #                .values()
        #                .annotate(employees=Count('employees_id'))
        #                .order_by()
        #                )
        #
        # results = Event.objects.raw(
        #     'SELECT ie.employees_id, ie.* FROM invoices_event ie order by ie.employees_id desc')
        # object_list = {}
        # for r in raw_list:
        #     if r.employees:
        #         l = object_list.get(r.employees.id)
        #         if l:
        #             l.append(r)
        #             # dd[r.employees.id] = l
        #         else:
        #             object_list[r.employees.id] = [r]

        extra_context = {'object_list': raw_list, 'root_url': config.ROOT_URL, 'form': self.form}

        return super(EventWeekListAdmin, self).changelist_view(request, extra_context)

    def get_form(self, request, obj=None, change=False, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if not request.user.is_superuser or not request.user.groups.filter(name="planning manager").exists():
            if len(form.base_fields) > 0:
                form.base_fields["event_report"].required = True
                form.base_fields["state"].choices = (3, _('Done')), (5, _('Not Done'))

        class ModelFormWithRequest(form):
            def __new__(cls, *args, **kwargs):
                kwargs['request'] = request
                return form(*args, **kwargs)

        return ModelFormWithRequest

    def get_readonly_fields(self, request, obj=None):
        if obj is not None:
            if not request.user.is_superuser:
                fs = [field.name for field in Event._meta.fields if field.name != "id"]
                if obj.employees.user.id == request.user.id:
                    return [f for f in fs if f not in ['event_report', 'state']]
                else:
                    return fs
        return self.readonly_fields


@admin.register(Alert)
class AlertAdmin(admin.ModelAdmin):
    list_display = (
        'text_alert', 'alert_level', 'date_alert', 'is_read', 'date_read', 'is_active', 'link_to_object', 'user')
    list_filter = ('alert_level', 'is_read', 'is_active', 'user')
    search_fields = ('text_alert', 'user__username')
    readonly_fields = ('date_alert', 'date_read', 'alert_created_by')

    # if not superuser show only alerts assigned by user
    def get_queryset(self, request):
        qs = super(AlertAdmin, self).get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.filter(user=request.user, is_active=True)

    # non superuser can only modify alerts assigned to him and only field is_read
    def get_readonly_fields(self, request, obj=None):
        if not request.user.is_superuser:
            return [f.name for f in self.model._meta.fields if f.name not in ['is_read', 'comment']]
        return self.readonly_fields


@admin.register(XeroToken)
class XeroTokenAdmin(admin.ModelAdmin):
    list_filter = ('access_token', 'expires_at',)


@admin.register(ConvadisOAuth2Token)
class ConvadisOAuth2TokenAdmin(admin.ModelAdmin):
    pass


@admin.register(EmployeeProxy)
class EmployeeProxyAdmin(admin.ModelAdmin):
    list_display = ['user']
    # all fields are read-only
    readonly_fields = [f.name for f in Employee._meta.fields]

    def get_queryset(self, request):
        """
        Filter the queryset to only include the employee data for the logged-in user.
        """
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs  # Superuser sees all records
        return qs.filter(user=request.user)

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        # Optionally, restrict the ability to change records
        return request.user.is_superuser or (obj is not None and obj.user == request.user)


@admin.register(DistanceMatrix)
class DistanceMatrixAdmin(admin.ModelAdmin):
    list_display = ['patient_origin', 'patient_destination', 'distance_in_km', 'duration_in_mn', 'created_at',
                    'updated_at']
    list_filter = (DistanceMatrixSmartPatientFilter,)
    search_fields = ('origin', 'destination')
    readonly_fields = ('created_at', 'updated_at')

    def changelist_view(self, request, extra_context=None):
        changelist = ChangeList(request,
                                self.model,
                                list(self.get_list_display(request)),
                                self.get_list_display_links(request, self.get_list_display(request)),
                                self.get_list_filter(request),
                                self.date_hierarchy,
                                self.get_search_fields(request),
                                self.list_select_related,
                                self.list_per_page,
                                self.list_max_show_all,
                                self.list_editable,
                                self,
                                self.sortable_by,
                                self.get_search_results)

        # Get a distinct list of patients for the filtered queryset
        patients = list(set(changelist.get_queryset(request).values_list('patient_origin__id', 'patient_origin__name')))
        # Attach it to request
        request.dynamic_patient_choices = [(str(patient_id), patient_name) for patient_id, patient_name in patients]

        return super().changelist_view(request, extra_context)


@admin.register(EmployeePaySlip)
class EmployeePaySlipAdmin(admin.ModelAdmin):
    list_display = ['employee', 'month', 'year', 'file']
    list_filter = ('employee', 'month', 'year')
    readonly_fields = ('created_at', 'updated_at')
    actions = ['send_payslip_by_email_in_attachment']

    def get_queryset(self, request):
        """
        Filter the queryset to only include the employee data for the logged-in user.
        """
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
    def send_payslip_by_email_in_attachment(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request, "Vous n'avez pas le droit d'envoyer des fiches de paie.", level=messages.WARNING)
            return
        payslip_sending_status = {}
        for payslip in queryset:
            payslip_sending_status[payslip.employee.user.username] = payslip.send_payslip_by_email_in_attachment()
        self.message_user(request, "Fiches de paie envoyées par email : %s" % payslip_sending_status)


# EmployeesMonthlyPayslipFile admin class
@admin.register(EmployeesMonthlyPayslipFile)
class EmployeesMonthlyPayslipFileAdmin(admin.ModelAdmin):
    list_display = ['month', 'year', 'created_at', 'updated_at']
    list_filter = ('month', 'year')
    readonly_fields = ('created_at', 'updated_at')

    def has_add_permission(self, request):
        return request.user.is_superuser

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser

    def has_change_permission(self, request, obj=None):
        return request.user.is_superuser

    def get_queryset(self, request):
        """
        Filter the queryset to only include the employee data for the logged-in user.
        """
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs


@admin.register(EmployeeVisit)
class EmployeeVisitAdmin(admin.ModelAdmin):
    list_filter = (SmartUserFilterForVisits, SmarPatientFilterForVisits)
    list_display = ('user', 'arrival_date_time', 'departure_date_time','display_url_on_google_maps',
                    'calcuate_duration_of_visit', 'patient')
    search_fields = ('patient', 'user')
    readonly_fields = ('created_at', 'updated_at', 'get_url_on_google_maps')
    actions = ['check_patient_addresses']

    def display_url_on_google_maps(self, obj):
        # return a url to google maps with the patient address
        return mark_safe(f'<a href="{obj.get_url_on_google_maps}" target="_blank">{obj.id}</a>')

    def check_patient_addresses(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request, "Vous n'avez pas le droit de vérifier les adresses des patients.", level=messages.WARNING)
            return
        visits_patients_found_dict = {}
        for visitq in queryset:
            visits_patients_found_dict[visitq] = visitq.check_if_address_is_known(visitq)
        self.message_user(request, "Adresses des patients vérifiées : %s" % visits_patients_found_dict)


